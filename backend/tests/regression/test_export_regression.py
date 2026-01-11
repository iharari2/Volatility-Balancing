#!/usr/bin/env python3
"""
Comprehensive regression tests for all export functionality.
These tests ensure that all export endpoints work correctly and don't break over time.
"""

import pytest
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

# Import the FastAPI app factory
from app.main import create_app

# Import domain entities for creating test data
from domain.entities.position import Position
from domain.entities.simulation_result import SimulationResult
from domain.value_objects.guardrails import GuardrailPolicy
from domain.value_objects.order_policy import OrderPolicy

# Import services for direct testing
from application.services.excel_export_service import ExcelExportService
from application.services.excel_template_service import ExcelTemplateService
from application.services.comprehensive_excel_export_service import ComprehensiveExcelExportService
from application.services.enhanced_excel_export_service import EnhancedExcelExportService

pytestmark = pytest.mark.online


class TestExportRegression:
    """Comprehensive regression tests for all export functionality"""

    @pytest.fixture
    def client(self):
        """Create test client"""
        app = create_app(enable_trading_worker=False)
        with TestClient(app) as test_client:
            yield test_client

    @pytest.fixture
    def mock_position(self):
        """Create a mock position for testing"""
        return Position(
            id="test_pos_123",
            tenant_id="default",
            portfolio_id="test_portfolio",
            asset_symbol="AAPL",
            qty=100.0,
            anchor_price=150.0,
            guardrails=GuardrailPolicy(min_stock_alloc_pct=0.25, max_stock_alloc_pct=0.75),
            order_policy=OrderPolicy(trigger_threshold_pct=0.03, rebalance_ratio=1.66667),
        )

    @pytest.fixture
    def mock_simulation_result(self):
        """Create a mock simulation result for testing"""
        return SimulationResult.create(
            ticker="AAPL",
            start_date="2024-01-01",
            end_date="2024-12-31",
            parameters={"trigger_threshold": 0.03, "rebalance_threshold": 0.05},
            metrics={
                "sharpe_ratio": 1.5,
                "return_pct": 15.2,
                "volatility": 0.12,
                "algorithm_pnl": 15000.0,
                "buy_hold_pnl": 12000.0,
                "max_drawdown": 0.08,
            },
            raw_data={
                "trade_log": [],
                "time_series_data": [],
                "total_trading_days": 252,
                "initial_cash": 100000.0,
                "algorithm_trades": 15,
            },
        )

    def _validate_excel_file(self, content: bytes, expected_sheets: list = None) -> bool:
        """Validate that content is a valid Excel file with expected structure"""
        try:
            # Load the Excel file from bytes
            wb = load_workbook(BytesIO(content))

            # Check that it has worksheets
            assert len(wb.worksheets) > 0, "Excel file should have at least one worksheet"

            # Check specific sheets if provided
            if expected_sheets:
                sheet_names = wb.sheetnames
                for expected_sheet in expected_sheets:
                    assert (
                        expected_sheet in sheet_names
                    ), f"Expected sheet '{expected_sheet}' not found"

            # Check that sheets have data
            for sheet in wb.worksheets:
                assert sheet.max_row > 1, f"Sheet '{sheet.title}' should have data beyond headers"
                assert sheet.max_column > 0, f"Sheet '{sheet.title}' should have columns"

            return True
        except Exception as e:
            pytest.fail(f"Excel validation failed: {e}")

    # API Endpoint Tests
    def test_trading_export_endpoint(self, client):
        """Test trading export API endpoint"""
        response = client.get("/v1/excel/trading/export")

        assert response.status_code == 200
        assert "spreadsheet" in response.headers.get("content-type", "")
        self._validate_excel_file(response.content)

    def test_positions_export_endpoint(self, client):
        """Test all positions export API endpoint"""
        response = client.get("/v1/excel/positions/export")

        assert response.status_code == 200
        assert "spreadsheet" in response.headers.get("content-type", "")
        self._validate_excel_file(response.content)

    def test_individual_position_export_endpoint(self, client):
        """Test individual position export API endpoint"""
        # Test with various tickers
        test_tickers = ["AAPL", "MSFT", "GOOGL", "ZIM", "BRK.A", "SPY"]

        for ticker in test_tickers:
            response = client.get(f"/v1/excel/positions/test_pos_{ticker}/export?ticker={ticker}")

            assert response.status_code == 200, f"Position export failed for {ticker}"
            assert "spreadsheet" in response.headers.get("content-type", "")
            self._validate_excel_file(response.content)

    def test_comprehensive_position_export_endpoint(self, client):
        """Test comprehensive position export API endpoint"""
        # Test with various tickers
        test_tickers = ["AAPL", "NVDA", "ZIM"]

        for ticker in test_tickers:
            response = client.get(
                f"/v1/excel/positions/test_pos_{ticker}/comprehensive-export?ticker={ticker}"
            )

            assert response.status_code == 200, f"Comprehensive export failed for {ticker}"
            assert "spreadsheet" in response.headers.get("content-type", "")
            self._validate_excel_file(
                response.content,
                expected_sheets=[
                    "Market Data",
                    "Position Data",
                    "Algorithm Data",
                    "Transaction Data",
                ],
            )

    def test_simulation_export_endpoint(self, client):
        """Test simulation export API endpoint"""
        test_tickers = ["AAPL", "ZIM", "SPY"]

        for ticker in test_tickers:
            response = client.get(f"/v1/excel/simulation/test_sim_{ticker}/export?ticker={ticker}")

            assert response.status_code == 200, f"Simulation export failed for {ticker}"
            assert "spreadsheet" in response.headers.get("content-type", "")
            self._validate_excel_file(response.content)

    def test_enhanced_simulation_export_endpoint(self, client):
        """Test enhanced simulation export API endpoint"""
        test_tickers = ["AAPL", "ZIM", "MSFT"]

        for ticker in test_tickers:
            response = client.get(
                f"/v1/excel/simulation/test_sim_{ticker}/enhanced-export?ticker={ticker}"
            )

            assert response.status_code == 200, f"Enhanced simulation export failed for {ticker}"
            assert "spreadsheet" in response.headers.get("content-type", "")
            self._validate_excel_file(
                response.content,
                expected_sheets=[
                    "Executive Summary",
                    "Comprehensive Time Series",
                    "Accounting Audit Trail",
                ],
            )

    # Service Layer Tests
    def test_excel_export_service(self, mock_position):
        """Test ExcelExportService directly"""
        service = ExcelExportService()

        # Test various export methods
        # Note: cash is no longer part of Position entity - it's in PortfolioCash
        # For export purposes, we can use a default cash value or omit it
        positions_data = [
            {
                "id": mock_position.id,
                "ticker": mock_position.ticker,
                "qty": mock_position.qty,
                "cash": getattr(mock_position, "cash", 5000.0),  # Cash from Position entity
                "anchor_price": mock_position.anchor_price,
            }
        ]

        # Test trading data export
        excel_data = service.export_trading_data(positions_data, [], [], "Test Trading Export")
        assert isinstance(excel_data, bytes)
        assert len(excel_data) > 1000  # Should be substantial
        self._validate_excel_file(excel_data)

    def test_excel_template_service(self, mock_simulation_result):
        """Test ExcelTemplateService directly"""
        service = ExcelTemplateService()

        # Test simulation results export
        excel_data = service.create_simulation_report(
            mock_simulation_result, mock_simulation_result.ticker
        )
        assert isinstance(excel_data, bytes)
        assert len(excel_data) > 1000
        self._validate_excel_file(excel_data)

    def test_comprehensive_excel_export_service(self, mock_position):
        """Test ComprehensiveExcelExportService directly"""
        service = ComprehensiveExcelExportService()

        # Test comprehensive position export
        excel_data = service.export_position_comprehensive_data(mock_position)
        assert isinstance(excel_data, bytes)
        assert len(excel_data) > 5000  # Should be comprehensive
        self._validate_excel_file(excel_data)

    def test_enhanced_excel_export_service(self, mock_simulation_result):
        """Test EnhancedExcelExportService directly"""
        service = EnhancedExcelExportService()

        # Test enhanced simulation export
        excel_data = service.export_comprehensive_simulation_report(
            mock_simulation_result, mock_simulation_result.ticker
        )
        assert isinstance(excel_data, bytes)
        assert len(excel_data) > 5000
        self._validate_excel_file(excel_data)

    # Error Handling Tests
    def test_invalid_ticker_handling(self, client):
        """Test that invalid tickers are handled gracefully"""
        invalid_tickers = ["INVALID123", "BADTICKER", ""]

        for ticker in invalid_tickers:
            if ticker:  # Skip empty ticker for this test
                response = client.get(f"/v1/excel/positions/test_pos/export?ticker={ticker}")
                # Should either succeed with fallback data or return a proper error
                assert response.status_code in [200, 400, 404, 500]

    def test_missing_simulation_data_handling(self, client):
        """Test handling of missing simulation data"""
        response = client.get("/v1/excel/simulation/nonexistent_sim/export?ticker=AAPL")
        # Should handle gracefully with mock data or proper error
        assert response.status_code in [200, 404]

    def test_malformed_position_id_handling(self, client):
        """Test handling of malformed position IDs"""
        malformed_ids = ["", "invalid-id", "pos_with_special_chars!@#"]

        for pos_id in malformed_ids:
            if pos_id:  # Skip empty ID
                response = client.get(f"/v1/excel/positions/{pos_id}/export")
                assert response.status_code in [200, 400, 404, 422]

    # Data Integrity Tests
    def test_export_data_consistency(self, client):
        """Test that exported data is consistent across different export formats"""
        ticker = "AAPL"
        pos_id = "test_consistency"

        # Get both basic and comprehensive exports for the same position
        basic_response = client.get(f"/v1/excel/positions/{pos_id}/export?ticker={ticker}")
        comprehensive_response = client.get(
            f"/v1/excel/positions/{pos_id}/comprehensive-export?ticker={ticker}"
        )

        assert basic_response.status_code == 200
        assert comprehensive_response.status_code == 200

        # Both should be valid Excel files
        self._validate_excel_file(basic_response.content)
        self._validate_excel_file(comprehensive_response.content)

        # Comprehensive should be larger (more data)
        assert len(comprehensive_response.content) >= len(basic_response.content)

    def test_ticker_parameter_propagation(self, client):
        """Test that ticker parameters are properly propagated through the system"""
        test_cases = [
            ("AAPL", "Apple Inc"),
            ("ZIM", "ZIM Integrated Shipping"),
            ("BRK.A", "Berkshire Hathaway Class A"),
            ("SPY", "SPDR S&P 500 ETF"),
        ]

        for ticker, description in test_cases:
            response = client.get(f"/v1/excel/positions/test_ticker_prop/export?ticker={ticker}")

            assert (
                response.status_code == 200
            ), f"Ticker propagation failed for {ticker} ({description})"
            self._validate_excel_file(response.content)

    # Performance Tests
    def test_export_performance(self, client):
        """Test that exports complete within reasonable time"""
        import time

        start_time = time.time()
        response = client.get("/v1/excel/positions/test_perf/comprehensive-export?ticker=AAPL")
        end_time = time.time()

        assert response.status_code == 200

        # Should complete within 30 seconds even with real data fetching
        execution_time = end_time - start_time
        assert execution_time < 30, f"Export took too long: {execution_time:.2f} seconds"

    def test_concurrent_exports(self, client):
        """Test that multiple exports can run concurrently"""
        import threading

        results = []

        def export_test(ticker, index):
            response = client.get(
                f"/v1/excel/positions/concurrent_test_{index}/export?ticker={ticker}"
            )
            results.append(
                (
                    index,
                    response.status_code,
                    len(response.content) if response.status_code == 200 else 0,
                )
            )

        # Start multiple exports concurrently
        threads = []
        tickers = ["AAPL", "MSFT", "GOOGL", "AMZN", "TSLA"]

        for i, ticker in enumerate(tickers):
            thread = threading.Thread(target=export_test, args=(ticker, i))
            threads.append(thread)
            thread.start()

        # Wait for all to complete
        for thread in threads:
            thread.join(timeout=60)  # 60 second timeout

        # All should succeed
        assert len(results) == len(tickers), "Not all concurrent exports completed"
        for index, status_code, content_length in results:
            assert status_code == 200, f"Concurrent export {index} failed with status {status_code}"
            assert content_length > 1000, f"Concurrent export {index} returned insufficient data"

    # File Format Validation Tests
    def test_excel_file_structure(self, client):
        """Test that exported Excel files have proper structure"""
        response = client.get("/v1/excel/positions/structure_test/comprehensive-export?ticker=AAPL")

        assert response.status_code == 200

        # Load and inspect the Excel file
        wb = load_workbook(BytesIO(response.content))

        # Should have multiple sheets
        assert len(wb.sheetnames) >= 4, f"Expected at least 4 sheets, got {len(wb.sheetnames)}"

        # Check specific sheet requirements
        required_sheets = ["Market Data", "Position Data", "Algorithm Data", "Transaction Data"]
        for sheet_name in required_sheets:
            assert sheet_name in wb.sheetnames, f"Required sheet '{sheet_name}' missing"

            ws = wb[sheet_name]
            # Should have headers
            assert ws.max_row >= 2, f"Sheet '{sheet_name}' should have data rows"
            assert ws.max_column >= 2, f"Sheet '{sheet_name}' should have multiple columns"

    def test_merged_cell_handling(self, client):
        """Test that merged cells don't cause errors (regression test for MergedCell issue)"""
        # This specifically tests the fix for the 'MergedCell' object has no attribute 'column_letter' error
        response = client.get("/v1/excel/simulation/merged_cell_test/enhanced-export?ticker=AAPL")

        assert response.status_code == 200, "Enhanced simulation export should handle merged cells"
        self._validate_excel_file(response.content)

        # Load and check that file doesn't have corrupted structure
        wb = load_workbook(BytesIO(response.content))
        for sheet in wb.worksheets:
            # Should be able to iterate through all cells without errors
            for row in sheet.iter_rows():
                for cell in row:
                    # Just accessing the cell should not raise an error
                    _ = cell.value


if __name__ == "__main__":
    # Run the tests
    pytest.main([__file__, "-v", "--tb=short"])
