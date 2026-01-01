#!/usr/bin/env python3
"""
Unit tests specifically for the MergedCell fix
Tests that the _safe_auto_fit_columns method properly handles merged cells
"""

import pytest
from openpyxl import Workbook

# Note: MergedCell is an internal class in openpyxl and not directly importable from utils.cell
# The test verifies that merged cells don't cause errors, not that we can import MergedCell
# If needed, import from openpyxl.cell.cell, but it's not used in this test

from application.services.comprehensive_excel_export_service import ComprehensiveExcelExportService
from application.services.excel_template_service import ExcelTemplateService
from domain.entities.position import Position
from domain.value_objects.guardrails import GuardrailPolicy
from domain.value_objects.order_policy import OrderPolicy


class TestMergedCellFix:
    """Test the fix for MergedCell column_letter error"""

    @pytest.fixture
    def mock_position(self):
        """Create a mock position for testing"""
        return Position(
            id="test_pos_123",
            tenant_id="default",
            portfolio_id="test_portfolio",
            asset_symbol="AAPL",
            qty=100.0,
            anchor_price=150.0,
            guardrails=GuardrailPolicy(min_stock_alloc_pct=0.25, max_stock_alloc_pct=0.75),
            order_policy=OrderPolicy(trigger_threshold_pct=0.03, rebalance_ratio=1.66667),
        )

    def test_safe_auto_fit_columns_with_merged_cells(self):
        """Test that _safe_auto_fit_columns handles merged cells without error"""
        service = ComprehensiveExcelExportService()

        # Create a workbook with merged cells
        wb = Workbook()
        ws = wb.active

        # Add some data
        ws["A1"] = "Header 1"
        ws["B1"] = "Header 2"
        ws["C1"] = "Header 3"

        ws["A2"] = "Data 1"
        ws["B2"] = "Data 2"
        ws["C2"] = "Data 3"

        # Merge some cells (this creates MergedCell objects)
        ws.merge_cells("A3:B3")
        ws["A3"] = "Merged Cell Content"

        # Add more data after merged cells
        ws["A4"] = "More data"
        ws["B4"] = "Even more data"
        ws["C4"] = "Last data"

        # This should NOT raise an error
        try:
            service._safe_auto_fit_columns(ws)
            success = True
        except AttributeError as e:
            if "'MergedCell' object has no attribute 'column_letter'" in str(e):
                success = False
                pytest.fail("MergedCell error not properly handled")
            else:
                raise
        except Exception as e:
            pytest.fail(f"Unexpected error in _safe_auto_fit_columns: {e}")

        assert success, "_safe_auto_fit_columns should handle merged cells"

        # Verify that column widths were actually set
        assert ws.column_dimensions["A"].width > 0
        assert ws.column_dimensions["B"].width > 0
        assert ws.column_dimensions["C"].width > 0

    def test_safe_auto_fit_columns_without_merged_cells(self):
        """Test that _safe_auto_fit_columns works normally without merged cells"""
        service = ComprehensiveExcelExportService()

        # Create a normal workbook without merged cells
        wb = Workbook()
        ws = wb.active

        # Add some data
        ws["A1"] = "Short"
        ws["B1"] = "Medium Length Header"
        ws["C1"] = "Very Long Header Content That Should Affect Column Width"

        ws["A2"] = "Data"
        ws["B2"] = "More Data"
        ws["C2"] = "Short"

        # This should work without issues
        service._safe_auto_fit_columns(ws)

        # Verify that column widths were set appropriately
        # Column C should be wider due to longer content
        assert ws.column_dimensions["C"].width > ws.column_dimensions["A"].width

    def test_excel_template_service_safe_auto_fit(self):
        """Test that ExcelTemplateService also has the safe auto-fit fix"""
        service = ExcelTemplateService()

        # Create a workbook with merged cells
        wb = Workbook()
        ws = wb.active

        # Add data and merge cells
        ws["A1"] = "Header"
        ws["B1"] = "Data"
        ws.merge_cells("A2:B2")
        ws["A2"] = "Merged Content"

        # Should not raise MergedCell error
        try:
            service._safe_auto_fit_columns(ws)
            success = True
        except AttributeError as e:
            if "'MergedCell' object has no attribute 'column_letter'" in str(e):
                success = False
                pytest.fail("ExcelTemplateService MergedCell error not properly handled")
            else:
                raise

        assert success

    def test_comprehensive_export_with_merged_cells(self, mock_position):
        """Test that full comprehensive export handles merged cells properly"""
        service = ComprehensiveExcelExportService()

        # This should complete without MergedCell errors
        try:
            excel_data = service.export_position_comprehensive_data(mock_position)
            assert isinstance(excel_data, bytes)
            assert len(excel_data) > 1000  # Should generate substantial content
        except AttributeError as e:
            if "'MergedCell' object has no attribute 'column_letter'" in str(e):
                pytest.fail("Comprehensive export failed due to MergedCell error")
            else:
                raise

    def test_empty_worksheet_handling(self):
        """Test that safe auto-fit handles empty worksheets gracefully"""
        service = ComprehensiveExcelExportService()

        # Create an empty worksheet
        wb = Workbook()
        ws = wb.active

        # Should handle empty worksheet without error
        try:
            service._safe_auto_fit_columns(ws)
            success = True
        except Exception as e:
            pytest.fail(f"Safe auto-fit failed on empty worksheet: {e}")

        assert success

    def test_worksheet_with_only_merged_cells(self):
        """Test handling of worksheet with only merged cells"""
        service = ComprehensiveExcelExportService()

        # Create worksheet with only merged cells
        wb = Workbook()
        ws = wb.active

        ws.merge_cells("A1:B2")
        ws["A1"] = "Only merged content"

        # Should handle this edge case
        try:
            service._safe_auto_fit_columns(ws)
            success = True
        except Exception as e:
            pytest.fail(f"Safe auto-fit failed on worksheet with only merged cells: {e}")

        assert success


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
