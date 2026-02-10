# =========================
# backend/tests/unit/application/test_activity_report_export.py
# =========================
"""Unit tests for ActivityReportExportService."""

import pytest
from datetime import datetime, timezone, timedelta
from io import BytesIO
from unittest.mock import MagicMock, patch
from openpyxl import load_workbook

from application.services.activity_report_export_service import ActivityReportExportService
from domain.entities.trade import Trade
from domain.entities.event import Event
from domain.entities.position import Position


@pytest.fixture
def mock_trades_repo():
    """Create mock trades repository."""
    repo = MagicMock()
    return repo


@pytest.fixture
def mock_events_repo():
    """Create mock events repository."""
    repo = MagicMock()
    return repo


@pytest.fixture
def mock_timeline_repo():
    """Create mock timeline repository."""
    repo = MagicMock()
    return repo


@pytest.fixture
def mock_positions_repo():
    """Create mock positions repository."""
    repo = MagicMock()
    return repo


@pytest.fixture
def sample_trades():
    """Create sample trades for testing."""
    base_time = datetime(2024, 1, 15, 10, 0, 0, tzinfo=timezone.utc)
    return [
        Trade(
            id="trade-001",
            tenant_id="default",
            portfolio_id="portfolio-1",
            order_id="order-001",
            position_id="position-1",
            side="BUY",
            qty=100.0,
            price=150.0,
            commission=1.50,
            executed_at=base_time,
        ),
        Trade(
            id="trade-002",
            tenant_id="default",
            portfolio_id="portfolio-1",
            order_id="order-002",
            position_id="position-1",
            side="SELL",
            qty=50.0,
            price=155.0,
            commission=0.78,
            executed_at=base_time + timedelta(days=5),
        ),
    ]


@pytest.fixture
def sample_events():
    """Create sample events for testing."""
    base_time = datetime(2024, 1, 15, 10, 0, 0, tzinfo=timezone.utc)
    return [
        Event(
            id="event-001",
            position_id="position-1",
            type="TRIGGER_EVALUATED",
            inputs={"price": 150.0, "anchor": 145.0},
            outputs={"trigger_fired": True, "direction": "UP"},
            message="Price triggered buy signal",
            ts=base_time - timedelta(minutes=5),
        ),
        Event(
            id="event-002",
            position_id="position-1",
            type="ORDER_SUBMITTED",
            inputs={"qty": 100, "side": "BUY"},
            outputs={"order_id": "order-001"},
            message="Buy order submitted",
            ts=base_time - timedelta(minutes=2),
        ),
    ]


@pytest.fixture
def sample_timeline():
    """Create sample timeline entries for testing."""
    base_time = datetime(2024, 1, 15, 10, 0, 0, tzinfo=timezone.utc)
    return [
        {
            "timestamp": base_time,
            "mode": "LIVE",
            "symbol": "AAPL",
            "position_id": "position-1",
            "effective_price": 150.0,
            "position_qty_before": 0.0,
            "position_cash_before": 15000.0,
            "position_total_value_before": 15000.0,
            "stock_pct": 0.0,
            "anchor_price": 145.0,
            "pct_change_from_anchor": 3.45,
            "trigger_fired": True,
            "trigger_direction": "UP",
            "action": "BUY",
            "action_reason": "Trigger fired above threshold",
            "trade_intent_qty": 100.0,
            "trade_id": "trade-001",
            "execution_price": 150.0,
            "execution_qty": 100.0,
            "position_qty_after": 100.0,
            "position_cash_after": 0.0,
            "position_total_value_after": 15000.0,
        },
        {
            "timestamp": base_time + timedelta(days=5),
            "mode": "LIVE",
            "symbol": "AAPL",
            "position_id": "position-1",
            "effective_price": 155.0,
            "position_qty_before": 100.0,
            "position_cash_before": 0.0,
            "position_total_value_before": 15500.0,
            "stock_pct": 100.0,
            "anchor_price": 150.0,
            "pct_change_from_anchor": 3.33,
            "trigger_fired": True,
            "trigger_direction": "UP",
            "action": "SELL",
            "action_reason": "Trigger fired above threshold",
            "trade_intent_qty": 50.0,
            "trade_id": "trade-002",
            "execution_price": 155.0,
            "execution_qty": 50.0,
            "position_qty_after": 50.0,
            "position_cash_after": 7750.0,
            "position_total_value_after": 15500.0,
        },
    ]


@pytest.fixture
def sample_positions():
    """Create sample positions for testing."""
    return [
        Position(
            id="position-1",
            tenant_id="default",
            portfolio_id="portfolio-1",
            asset_symbol="AAPL",
            qty=50.0,
            cash=7750.0,
            anchor_price=150.0,
        ),
    ]


@pytest.fixture
def service(mock_trades_repo, mock_events_repo, mock_timeline_repo, mock_positions_repo):
    """Create service instance with mocked dependencies."""
    return ActivityReportExportService(
        trades_repo=mock_trades_repo,
        events_repo=mock_events_repo,
        timeline_repo=mock_timeline_repo,
        positions_repo=mock_positions_repo,
    )


class TestActivityReportExportService:
    """Tests for ActivityReportExportService."""

    def test_export_creates_workbook_with_all_sheets(
        self,
        service,
        mock_trades_repo,
        mock_events_repo,
        mock_timeline_repo,
        mock_positions_repo,
        sample_trades,
        sample_events,
        sample_timeline,
        sample_positions,
    ):
        """Test that export creates workbook with all required sheets."""
        # Setup mocks
        mock_trades_repo.list_by_portfolio.return_value = sample_trades
        mock_events_repo.list_for_position.return_value = sample_events
        mock_timeline_repo.list_by_portfolio.return_value = sample_timeline
        mock_positions_repo.list_all.return_value = sample_positions

        # Export
        excel_bytes = service.export_activity_report(
            tenant_id="default",
            portfolio_id="portfolio-1",
        )

        # Load workbook and check sheets
        wb = load_workbook(BytesIO(excel_bytes))
        sheet_names = wb.sheetnames

        assert "Summary" in sheet_names
        assert "Activity Log" in sheet_names
        assert "Trades" in sheet_names
        assert "Evaluation Timeline" in sheet_names
        assert "Dividends" in sheet_names

    def test_export_summary_sheet_has_metrics(
        self,
        service,
        mock_trades_repo,
        mock_events_repo,
        mock_timeline_repo,
        mock_positions_repo,
        sample_trades,
        sample_events,
        sample_timeline,
        sample_positions,
    ):
        """Test that summary sheet contains key metrics."""
        mock_trades_repo.list_by_portfolio.return_value = sample_trades
        mock_events_repo.list_for_position.return_value = sample_events
        mock_timeline_repo.list_by_portfolio.return_value = sample_timeline
        mock_positions_repo.list_all.return_value = sample_positions

        excel_bytes = service.export_activity_report(
            tenant_id="default",
            portfolio_id="portfolio-1",
        )

        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Summary"]

        # Extract all values from the sheet
        values = []
        for row in ws.iter_rows(values_only=True):
            values.extend([v for v in row if v])

        # Check for expected content
        assert "Total Trades" in values
        assert "Buy Trades" in values
        assert "Sell Trades" in values
        assert "Total Commission" in values
        assert "Total Dividends" in values

    def test_export_trades_sheet_has_trade_data(
        self,
        service,
        mock_trades_repo,
        mock_events_repo,
        mock_timeline_repo,
        mock_positions_repo,
        sample_trades,
        sample_events,
        sample_timeline,
        sample_positions,
    ):
        """Test that trades sheet contains trade data."""
        mock_trades_repo.list_by_portfolio.return_value = sample_trades
        mock_events_repo.list_for_position.return_value = sample_events
        mock_timeline_repo.list_by_portfolio.return_value = sample_timeline
        mock_positions_repo.list_all.return_value = sample_positions

        excel_bytes = service.export_activity_report(
            tenant_id="default",
            portfolio_id="portfolio-1",
        )

        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Trades"]

        # Check header row
        headers = [cell.value for cell in ws[1]]
        assert "executed_at" in headers
        assert "side" in headers
        assert "qty" in headers
        assert "price" in headers
        assert "commission" in headers

        # Check data rows (should have 2 trades + header)
        row_count = sum(1 for _ in ws.iter_rows())
        assert row_count >= 3  # Header + 2 trades

    def test_activity_log_merges_sources_chronologically(
        self,
        service,
        mock_trades_repo,
        mock_events_repo,
        mock_timeline_repo,
        mock_positions_repo,
        sample_trades,
        sample_events,
        sample_timeline,
        sample_positions,
    ):
        """Test that activity log merges all sources chronologically."""
        mock_trades_repo.list_by_portfolio.return_value = sample_trades
        mock_events_repo.list_for_position.return_value = sample_events
        mock_timeline_repo.list_by_portfolio.return_value = sample_timeline
        mock_positions_repo.list_all.return_value = sample_positions

        excel_bytes = service.export_activity_report(
            tenant_id="default",
            portfolio_id="portfolio-1",
        )

        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Activity Log"]

        # Check that we have multiple event types
        event_types = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1]:  # event_type column
                event_types.add(row[1])

        # Should have TRADE and other event types
        assert "TRADE" in event_types

    def test_export_with_date_filter(
        self,
        service,
        mock_trades_repo,
        mock_events_repo,
        mock_timeline_repo,
        mock_positions_repo,
        sample_positions,
    ):
        """Test that date filtering is passed to repositories."""
        mock_trades_repo.list_by_portfolio.return_value = []
        mock_events_repo.list_for_position.return_value = []
        mock_timeline_repo.list_by_portfolio.return_value = []
        mock_positions_repo.list_all.return_value = sample_positions

        start_date = datetime(2024, 1, 1, tzinfo=timezone.utc)
        end_date = datetime(2024, 1, 31, tzinfo=timezone.utc)

        service.export_activity_report(
            tenant_id="default",
            portfolio_id="portfolio-1",
            start_date=start_date,
            end_date=end_date,
        )

        # Verify date filters were passed
        mock_trades_repo.list_by_portfolio.assert_called_once()
        call_kwargs = mock_trades_repo.list_by_portfolio.call_args.kwargs
        assert call_kwargs.get("start_date") == start_date
        assert call_kwargs.get("end_date") == end_date

    def test_export_with_position_filter(
        self,
        service,
        mock_trades_repo,
        mock_events_repo,
        mock_timeline_repo,
        mock_positions_repo,
    ):
        """Test that position filtering works correctly."""
        mock_trades_repo.list_by_portfolio.return_value = []
        mock_events_repo.list_for_position.return_value = []
        mock_timeline_repo.list_by_position.return_value = []

        service.export_activity_report(
            tenant_id="default",
            portfolio_id="portfolio-1",
            position_id="position-1",
        )

        # Should use list_by_position instead of list_by_portfolio for timeline
        mock_timeline_repo.list_by_position.assert_called_once()

    def test_export_with_empty_data(
        self,
        service,
        mock_trades_repo,
        mock_events_repo,
        mock_timeline_repo,
        mock_positions_repo,
    ):
        """Test that export works with empty data."""
        mock_trades_repo.list_by_portfolio.return_value = []
        mock_events_repo.list_for_position.return_value = []
        mock_timeline_repo.list_by_portfolio.return_value = []
        mock_positions_repo.list_all.return_value = []

        excel_bytes = service.export_activity_report(
            tenant_id="default",
            portfolio_id="portfolio-1",
        )

        # Should still create a valid workbook
        wb = load_workbook(BytesIO(excel_bytes))
        assert len(wb.sheetnames) == 5

    def test_dividends_sheet_filters_dividend_events(
        self,
        service,
        mock_trades_repo,
        mock_events_repo,
        mock_timeline_repo,
        mock_positions_repo,
    ):
        """Test that dividends sheet only includes dividend events."""
        dividend_timeline = [
            {
                "timestamp": datetime(2024, 1, 20, tzinfo=timezone.utc),
                "mode": "LIVE",
                "symbol": "AAPL",
                "position_id": "position-1",
                "dividend_applied": True,
                "dividend_ex_date": "2024-01-18",
                "dividend_pay_date": "2024-01-25",
                "dividend_rate": 0.24,
                "dividend_gross_value": 24.0,
                "dividend_tax": 6.0,
                "dividend_net_value": 18.0,
                "position_qty_before": 100.0,
                "position_cash_before": 1000.0,
                "position_cash_after": 1018.0,
            },
            {
                "timestamp": datetime(2024, 1, 15, tzinfo=timezone.utc),
                "mode": "LIVE",
                "symbol": "AAPL",
                "position_id": "position-1",
                "dividend_applied": False,
                "action": "HOLD",
            },
        ]

        mock_trades_repo.list_by_portfolio.return_value = []
        mock_events_repo.list_for_position.return_value = []
        mock_timeline_repo.list_by_portfolio.return_value = dividend_timeline
        mock_positions_repo.list_all.return_value = []

        excel_bytes = service.export_activity_report(
            tenant_id="default",
            portfolio_id="portfolio-1",
        )

        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Dividends"]

        # Count data rows (excluding header)
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        non_empty_rows = [r for r in data_rows if any(r)]

        # Should only have 1 dividend row
        assert len(non_empty_rows) == 1


class TestMergeToActivityLog:
    """Tests for the _merge_to_activity_log method."""

    def test_trades_are_converted_to_activity_entries(
        self,
        service,
        sample_trades,
    ):
        """Test that trades are properly converted to activity log entries."""
        activity_log = service._merge_to_activity_log(
            trades=sample_trades,
            events=[],
            timeline=[],
        )

        trade_entries = [e for e in activity_log if e["event_type"] == "TRADE"]
        assert len(trade_entries) == 2

        # Check first trade entry
        first_trade = trade_entries[0]
        assert first_trade["position_id"] == "position-1"
        assert "BUY" in first_trade["description"]
        assert first_trade["related_trade_id"] == "trade-001"

    def test_events_are_converted_to_activity_entries(
        self,
        service,
        sample_events,
    ):
        """Test that events are properly converted to activity log entries."""
        activity_log = service._merge_to_activity_log(
            trades=[],
            events=sample_events,
            timeline=[],
        )

        assert len(activity_log) == 2

        # Events should preserve their type
        event_types = {e["event_type"] for e in activity_log}
        assert "TRIGGER_EVALUATED" in event_types
        assert "ORDER_SUBMITTED" in event_types

    def test_activity_log_is_sorted_chronologically(
        self,
        service,
        sample_trades,
        sample_events,
        sample_timeline,
    ):
        """Test that activity log is sorted by timestamp."""
        activity_log = service._merge_to_activity_log(
            trades=sample_trades,
            events=sample_events,
            timeline=sample_timeline,
        )

        timestamps = [e["timestamp"] for e in activity_log if e["timestamp"]]

        # Verify sorted order
        for i in range(len(timestamps) - 1):
            assert timestamps[i] <= timestamps[i + 1], "Activity log should be chronologically sorted"
