# =========================
# backend/application/services/excel_export_service.py
# =========================

from datetime import datetime, timezone
from typing import Dict, List, Any, Optional
import pandas as pd
from openpyxl.styles import Font, PatternFill

from domain.entities.optimization_result import OptimizationResults
from domain.entities.simulation_result import SimulationResult
from domain.entities.position import Position
from application.services.excel_template_service import ExcelTemplateService
from application.services.enhanced_excel_export_service import EnhancedExcelExportService
from application.services.comprehensive_excel_export_service import ComprehensiveExcelExportService


class ExcelExportService:
    """Service for exporting data to Excel format with comprehensive reporting."""

    def __init__(self):
        """Initialize the Excel export service."""
        self.workbook = None
        self.ws = None

    def export_optimization_results(
        self, results: OptimizationResults, config_name: str = "Optimization Results"
    ) -> bytes:
        """Export optimization results to Excel format."""
        # Use the template service for professional formatting
        template_service = ExcelTemplateService()
        return template_service.create_optimization_report(results, config_name)

    def export_simulation_results(
        self, simulation_result: SimulationResult, ticker: str = "UNKNOWN"
    ) -> bytes:
        """Export simulation results to Excel format."""
        # Use the template service for professional formatting
        template_service = ExcelTemplateService()
        return template_service.create_simulation_report(simulation_result, ticker)

    def export_trading_data(
        self,
        positions_data: List[Dict[str, Any]],
        trades_data: List[Dict[str, Any]],
        orders_data: List[Dict[str, Any]],
        title: str = "Trading Data Export",
    ) -> bytes:
        """Export trading data (positions, trades, orders) to Excel format."""
        # Use the template service for professional formatting
        template_service = ExcelTemplateService()
        return template_service.create_trading_audit_report(
            positions_data, trades_data, orders_data, title
        )

    def export_activity_log(
        self,
        events: List[Dict[str, Any]],
        position_id: str,
        ticker: str = "UNKNOWN",
    ) -> bytes:
        """Export activity log (events) to Excel format."""
        template_service = ExcelTemplateService()
        return template_service.create_activity_log_report(events, position_id, ticker)

    def export_comprehensive_simulation_report(
        self, simulation_result: SimulationResult, ticker: str = "UNKNOWN"
    ) -> bytes:
        """Export comprehensive simulation report with complete data logs."""
        enhanced_service = EnhancedExcelExportService()
        return enhanced_service.export_comprehensive_simulation_report(simulation_result, ticker)

    def export_trading_audit_report_enhanced(
        self,
        positions_data: List[Dict[str, Any]],
        trades_data: List[Dict[str, Any]],
        orders_data: List[Dict[str, Any]],
        events_data: List[Dict[str, Any]],
        title: str = "Trading Audit Report",
    ) -> bytes:
        """Export comprehensive trading audit report with events."""
        enhanced_service = EnhancedExcelExportService()
        return enhanced_service.export_trading_audit_report(
            positions_data, trades_data, orders_data, events_data, title
        )

    def export_position_comprehensive_data(
        self,
        position: Position,
        market_data: List[Dict[str, Any]] = None,
        transaction_data: List[Dict[str, Any]] = None,
        simulation_data: Optional[SimulationResult] = None,
    ) -> bytes:
        """Export comprehensive per-position data to Excel."""
        comprehensive_service = ComprehensiveExcelExportService()
        return comprehensive_service.export_position_comprehensive_data(
            position, market_data, transaction_data, simulation_data
        )

    def _create_optimization_summary_sheet(self, results: OptimizationResults, config_name: str):
        """Create optimization summary sheet."""
        ws = self.workbook.create_sheet("Summary", 0)

        # Header
        ws["A1"] = f"{config_name} - Summary Report"
        ws["A1"].font = Font(size=16, bold=True)
        ws["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}"

        # Configuration info
        ws["A4"] = "Configuration ID:"
        ws["B4"] = str(results.config_id)
        ws["A5"] = "Total Results:"
        ws["B5"] = len(results.results)
        ws["A6"] = "Completed Results:"
        ws["B6"] = len(results.get_completed_results())
        ws["A7"] = "Failed Results:"
        ws["B7"] = len(results.get_failed_results())

        # Best result summary
        if results.best_result:
            ws["A9"] = "Best Result:"
            ws["A9"].font = Font(bold=True)
            ws["A10"] = "Combination ID:"
            ws["B10"] = results.best_result.combination_id
            ws["A11"] = "Parameters:"
            for i, (param, value) in enumerate(
                results.best_result.parameter_combination.parameters.items()
            ):
                ws[f"A{12+i}"] = f"  {param}:"
                ws[f"B{12+i}"] = str(value)

            # Metrics
            metrics_start_row = 12 + len(results.best_result.parameter_combination.parameters)
            ws[f"A{metrics_start_row}"] = "Metrics:"
            ws[f"A{metrics_start_row}"].font = Font(bold=True)
            for i, (metric, value) in enumerate(results.best_result.metrics.items()):
                ws[f"A{metrics_start_row+1+i}"] = f"  {metric.value}:"
                ws[f"B{metrics_start_row+1+i}"] = f"{value:.6f}"

    def _create_optimization_details_sheet(self, results: OptimizationResults):
        """Create detailed optimization results sheet."""
        ws = self.workbook.create_sheet("Detailed Results")

        if not results.results:
            ws["A1"] = "No results available"
            return

        # Prepare data for DataFrame
        data = []
        for result in results.results:
            row = {
                "ID": str(result.id),
                "Combination ID": result.combination_id,
                "Status": result.status.value,
                "Created At": result.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                "Completed At": (
                    result.completed_at.strftime("%Y-%m-%d %H:%M:%S") if result.completed_at else ""
                ),
                "Execution Time (s)": result.execution_time_seconds or "",
                "Error Message": result.error_message or "",
            }

            # Add parameters
            for param, value in result.parameter_combination.parameters.items():
                row[f"Param_{param}"] = value

            # Add metrics
            for metric, value in result.metrics.items():
                row[f"Metric_{metric.value}"] = value

            data.append(row)

        # Create DataFrame and write to Excel
        df = pd.DataFrame(data)

        # Write headers
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Write data
        for row_num, row_data in enumerate(df.itertuples(index=False), 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_parameter_analysis_sheet(self, results: OptimizationResults):
        """Create parameter analysis sheet."""
        ws = self.workbook.create_sheet("Parameter Analysis")

        if not results.results:
            ws["A1"] = "No results available"
            return

        # Get all unique parameters
        all_params = set()
        for result in results.results:
            all_params.update(result.parameter_combination.parameters.keys())

        # Create parameter statistics
        ws["A1"] = "Parameter Analysis"
        ws["A1"].font = Font(size=14, bold=True)

        row = 3
        for param in sorted(all_params):
            ws[f"A{row}"] = f"Parameter: {param}"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

            # Get all values for this parameter
            values = []
            for result in results.results:
                if param in result.parameter_combination.parameters:
                    values.append(result.parameter_combination.parameters[param])

            if values:
                # Calculate statistics
                if all(isinstance(v, (int, float)) for v in values):
                    ws[f"B{row}"] = "Min:"
                    ws[f"C{row}"] = min(values)
                    ws[f"D{row}"] = "Max:"
                    ws[f"E{row}"] = max(values)
                    ws[f"F{row}"] = "Mean:"
                    ws[f"G{row}"] = sum(values) / len(values)
                    row += 1

                ws[f"B{row}"] = "Unique Values:"
                ws[f"C{row}"] = len(set(values))
                ws[f"D{row}"] = "Total Count:"
                ws[f"E{row}"] = len(values)
                row += 2

    def _create_performance_metrics_sheet(self, results: OptimizationResults):
        """Create performance metrics sheet."""
        ws = self.workbook.create_sheet("Performance Metrics")

        if not results.results:
            ws["A1"] = "No results available"
            return

        # Get all unique metrics
        all_metrics = set()
        for result in results.results:
            all_metrics.update(result.metrics.keys())

        # Create metrics statistics
        ws["A1"] = "Performance Metrics Analysis"
        ws["A1"].font = Font(size=14, bold=True)

        row = 3
        for metric in sorted(all_metrics, key=lambda x: x.value):
            ws[f"A{row}"] = f"Metric: {metric.value}"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

            # Get all values for this metric
            values = []
            for result in results.results:
                if metric in result.metrics:
                    values.append(result.metrics[metric])

            if values:
                ws[f"B{row}"] = "Min:"
                ws[f"C{row}"] = f"{min(values):.6f}"
                ws[f"D{row}"] = "Max:"
                ws[f"E{row}"] = f"{max(values):.6f}"
                ws[f"F{row}"] = "Mean:"
                ws[f"G{row}"] = f"{sum(values) / len(values):.6f}"
                ws[f"H{row}"] = "Count:"
                ws[f"I{row}"] = len(values)
                row += 2

    def _create_simulation_summary_sheet(self, simulation_result: SimulationResult, ticker: str):
        """Create simulation summary sheet."""
        ws = self.workbook.create_sheet("Simulation Summary", 0)

        # Header
        ws["A1"] = f"Simulation Results - {ticker}"
        ws["A1"].font = Font(size=16, bold=True)
        ws["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}"

        # Simulation details
        ws["A4"] = "Ticker:"
        ws["B4"] = ticker
        ws["A5"] = "Start Date:"
        ws["B5"] = simulation_result.start_date
        ws["A6"] = "End Date:"
        ws["B6"] = simulation_result.end_date

        # Performance metrics
        ws["A8"] = "Performance Metrics:"
        ws["A8"].font = Font(bold=True)

        if hasattr(simulation_result, "metrics") and simulation_result.metrics:
            row = 9
            for metric, value in simulation_result.metrics.items():
                ws[f"A{row}"] = f"{metric}:"
                ws[f"B{row}"] = f"{value:.6f}"
                row += 1

    def _create_trade_log_sheet(self, trade_log: List[Dict[str, Any]]):
        """Create trade log sheet."""
        ws = self.workbook.create_sheet("Trade Log")

        if not trade_log:
            ws["A1"] = "No trade data available"
            return

        # Create DataFrame from trade log
        df = pd.DataFrame(trade_log)

        # Write headers
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Write data
        for row_num, row_data in enumerate(df.itertuples(index=False), 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_daily_returns_sheet(self, daily_returns: List[Dict[str, Any]]):
        """Create daily returns sheet."""
        ws = self.workbook.create_sheet("Daily Returns")

        if not daily_returns:
            ws["A1"] = "No daily returns data available"
            return

        # Create DataFrame from daily returns
        df = pd.DataFrame(daily_returns)

        # Write headers
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Write data
        for row_num, row_data in enumerate(df.itertuples(index=False), 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_price_data_sheet(self, price_data: List[Dict[str, Any]]):
        """Create price data sheet."""
        ws = self.workbook.create_sheet("Price Data")

        if not price_data:
            ws["A1"] = "No price data available"
            return

        # Create DataFrame from price data
        df = pd.DataFrame(price_data)

        # Write headers
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Write data
        for row_num, row_data in enumerate(df.itertuples(index=False), 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_positions_sheet(self, positions_data: List[Dict[str, Any]]):
        """Create positions sheet."""
        ws = self.workbook.create_sheet("Positions")

        if not positions_data:
            ws["A1"] = "No positions data available"
            return

        # Create DataFrame from positions data
        df = pd.DataFrame(positions_data)

        # Write headers
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Write data
        for row_num, row_data in enumerate(df.itertuples(index=False), 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_trades_sheet(self, trades_data: List[Dict[str, Any]]):
        """Create trades sheet."""
        ws = self.workbook.create_sheet("Trades")

        if not trades_data:
            ws["A1"] = "No trades data available"
            return

        # Create DataFrame from trades data
        df = pd.DataFrame(trades_data)

        # Write headers
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Write data
        for row_num, row_data in enumerate(df.itertuples(index=False), 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_orders_sheet(self, orders_data: List[Dict[str, Any]]):
        """Create orders sheet."""
        ws = self.workbook.create_sheet("Orders")

        if not orders_data:
            ws["A1"] = "No orders data available"
            return

        # Create DataFrame from orders data
        df = pd.DataFrame(orders_data)

        # Write headers
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Write data
        for row_num, row_data in enumerate(df.itertuples(index=False), 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_trading_summary_sheet(
        self,
        positions_data: List[Dict[str, Any]],
        trades_data: List[Dict[str, Any]],
        orders_data: List[Dict[str, Any]],
        title: str,
    ):
        """Create trading summary sheet."""
        ws = self.workbook.create_sheet("Summary", 0)

        # Header
        ws["A1"] = title
        ws["A1"].font = Font(size=16, bold=True)
        ws["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}"

        # Summary statistics
        ws["A4"] = "Summary Statistics:"
        ws["A4"].font = Font(bold=True)
        ws["A5"] = "Total Positions:"
        ws["B5"] = len(positions_data)
        ws["A6"] = "Total Trades:"
        ws["B6"] = len(trades_data)
        ws["A7"] = "Total Orders:"
        ws["B7"] = len(orders_data)
