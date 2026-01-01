# =========================
# backend/application/services/timeline_excel_export_service.py
# =========================
"""
Excel export service using the canonical PositionEvaluationTimeline table.

This service exports data according to the specification:
- Sheet 1: Timeline (chronological, non-verbose columns)
- Sheet 2: Timeline_Verbose (all columns)
- Sheet 3: Trades (rows where action in BUY/SELL)
- Sheet 4: Dividends (rows where dividend_applied = true)
- Sheet 5: Summary (KPIs, totals, drawdown, volatility)

All data comes from PositionEvaluationTimeline - no duplicated logic.
"""

from __future__ import annotations
from typing import Dict, Any, List, Optional
from datetime import datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from domain.ports.evaluation_timeline_repo import EvaluationTimelineRepo


class TimelineExcelExportService:
    """Excel export service using PositionEvaluationTimeline table."""

    def __init__(self, timeline_repo: EvaluationTimelineRepo):
        self.timeline_repo = timeline_repo

    def export_portfolio_timeline(
        self,
        tenant_id: str,
        portfolio_id: str,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        mode: Optional[str] = None,  # 'LIVE' or 'SIMULATION'
        output_path: Optional[str] = None,
    ) -> bytes:
        """
        Export portfolio timeline to Excel according to specification.

        Args:
            tenant_id: Tenant identifier
            portfolio_id: Portfolio identifier
            start_date: Optional start date filter
            end_date: Optional end date filter
            mode: Optional mode filter ('LIVE' or 'SIMULATION')
            output_path: Optional path to save file (if None, returns bytes)

        Returns:
            Excel file as bytes
        """
        # Get timeline data
        timeline_rows = self.timeline_repo.list_by_portfolio(
            tenant_id=tenant_id,
            portfolio_id=portfolio_id,
            mode=mode,
            start_date=start_date,
            end_date=end_date,
        )

        # Sort by timestamp (chronological)
        timeline_rows.sort(key=lambda x: x.get("timestamp") or datetime.min)

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Sheet 1: Timeline (non-verbose columns)
        self._create_timeline_sheet(wb, timeline_rows, verbose=False)

        # Sheet 2: Timeline_Verbose (all columns)
        self._create_timeline_sheet(wb, timeline_rows, verbose=True)

        # Sheet 3: Trades (action in BUY/SELL)
        self._create_trades_sheet(wb, timeline_rows)

        # Sheet 4: Dividends (dividend_applied = true)
        self._create_dividends_sheet(wb, timeline_rows)

        # Sheet 5: Summary (KPIs, totals, drawdown, volatility)
        self._create_summary_sheet(wb, timeline_rows, tenant_id, portfolio_id)

        # Save or return
        if output_path:
            wb.save(output_path)
            return b""
        else:
            from io import BytesIO

            buffer = BytesIO()
            wb.save(buffer)
            return buffer.getvalue()

    def _create_timeline_sheet(
        self, wb: Workbook, rows: List[Dict[str, Any]], verbose: bool = False
    ):
        """Create Timeline or Timeline_Verbose sheet."""
        ws = wb.create_sheet(title="Timeline_Verbose" if verbose else "Timeline")

        # Define columns based on verbose mode
        if verbose:
            # All columns from specification
            columns = [
                # Identification & context
                "timestamp",
                "mode",
                "tenant_id",
                "portfolio_id",
                "portfolio_name",
                "position_id",
                "symbol",
                "exchange",
                "market_session",
                "evaluation_seq",
                "trace_id",
                "source",
                # Market data - OHLCV
                "open_price",
                "high_price",
                "low_price",
                "close_price",
                "volume",
                # Market data - Quote details
                "last_trade_price",
                "best_bid",
                "best_ask",
                "official_close_price",
                "effective_price",
                "price_policy_requested",
                "price_policy_effective",
                "price_fallback_reason",
                "data_provider",
                "is_market_hours",
                "allow_after_hours",
                "trading_hours_policy",
                "price_validation_valid",
                "is_fresh",
                "is_inline",
                # Dividend data
                "dividend_declared",
                "dividend_ex_date",
                "dividend_pay_date",
                "dividend_rate",
                "dividend_gross_value",
                "dividend_tax",
                "dividend_net_value",
                "dividend_applied",
                # Position state before
                "position_qty_before",
                "position_cash_before",
                "position_stock_value_before",
                "position_total_value_before",
                "cash_pct",
                "stock_pct",
                "position_dividend_receivable_before",
                # Strategy state - Anchor
                "anchor_price",
                "pct_change_from_anchor",
                "pct_change_from_prev",
                "anchor_updated",
                "anchor_reset_old_value",
                "anchor_reset_reason",
                # Strategy state - Triggers
                "trigger_up_threshold",
                "trigger_down_threshold",
                "trigger_direction",
                "trigger_fired",
                "trigger_reason",
                # Strategy state - Guardrails
                "guardrail_min_stock_pct",
                "guardrail_max_stock_pct",
                "guardrail_max_trade_pct",
                "guardrail_max_orders_per_day",
                "guardrail_allowed",
                "guardrail_block_reason",
                # Order policy
                "order_policy_rebalance_ratio",
                "order_policy_commission_rate",
                "order_policy_min_qty",
                "order_policy_min_notional",
                # Action decision
                "action",
                "action_reason",
                "trade_intent_qty",
                "trade_intent_value",
                "trade_intent_cash_delta",
                # Execution result
                "order_id",
                "trade_id",
                "execution_price",
                "execution_qty",
                "commission_rate",
                "commission_value",
                "execution_status",
                "execution_timestamp",
                # Position state after
                "position_qty_after",
                "position_cash_after",
                "position_stock_value_after",
                "position_total_value_after",
                "position_stock_pct_after",
                "new_anchor_price",
                # Portfolio impact
                "portfolio_total_value_before",
                "portfolio_total_value_after",
                "portfolio_cash_before",
                "portfolio_cash_after",
                "portfolio_stock_value_before",
                "portfolio_stock_value_after",
                "position_weight_pct_before",
                "position_weight_pct_after",
                # Verbose explanation fields
                "evaluation_notes",
                "pricing_notes",
                "trigger_notes",
                "guardrail_notes",
                "action_notes",
                "warnings",
            ]
        else:
            # Compact columns (non-verbose)
            columns = [
                "timestamp",
                "mode",
                "symbol",
                "position_id",
                "effective_price",
                "is_market_hours",
                "allow_after_hours",
                "position_qty_before",
                "position_cash_before",
                "position_stock_value_before",
                "position_total_value_before",
                "stock_pct",
                "anchor_price",
                "pct_change_from_anchor",
                "trigger_fired",
                "trigger_direction",
                "trigger_reason",
                "guardrail_allowed",
                "guardrail_block_reason",
                "action",
                "action_reason",
                "trade_intent_qty",
                "trade_intent_value",
                "order_id",
                "trade_id",
                "execution_price",
                "execution_qty",
                "execution_status",
                "position_qty_after",
                "position_cash_after",
                "position_total_value_after",
                "portfolio_total_value_before",
                "portfolio_total_value_after",
                "position_weight_pct_before",
                "position_weight_pct_after",
            ]

        # Write header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data rows
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                value = row_data.get(col_name)
                # Format datetime values
                if isinstance(value, datetime):
                    value = value.isoformat()
                elif isinstance(value, Decimal):
                    value = float(value)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

    def _create_trades_sheet(self, wb: Workbook, rows: List[Dict[str, Any]]):
        """Create Trades sheet (rows where action in BUY/SELL)."""
        ws = wb.create_sheet(title="Trades")

        # Filter to trades only
        trade_rows = [r for r in rows if r.get("action") in ("BUY", "SELL")]

        columns = [
            "timestamp",
            "symbol",
            "action",
            "action_reason",
            "trade_intent_qty",
            "trade_intent_value",
            "trade_intent_cash_delta",
            "execution_price",
            "execution_qty",
            "execution_status",
            "commission_rate",
            "commission_value",
            "order_id",
            "trade_id",
            "position_qty_before",
            "position_qty_after",
            "position_cash_before",
            "position_cash_after",
            "position_total_value_before",
            "position_total_value_after",
        ]

        # Write header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font

        # Write data
        for row_idx, row_data in enumerate(trade_rows, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                value = row_data.get(col_name)
                if isinstance(value, datetime):
                    value = value.isoformat()
                elif isinstance(value, Decimal):
                    value = float(value)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

    def _create_dividends_sheet(self, wb: Workbook, rows: List[Dict[str, Any]]):
        """Create Dividends sheet (rows where dividend_applied = true)."""
        ws = wb.create_sheet(title="Dividends")

        # Filter to dividends only
        dividend_rows = [r for r in rows if r.get("dividend_applied", False)]

        columns = [
            "timestamp",
            "symbol",
            "position_id",
            "dividend_declared",
            "dividend_ex_date",
            "dividend_pay_date",
            "dividend_rate",
            "dividend_gross_value",
            "dividend_tax",
            "dividend_net_value",
            "position_qty_before",
            "position_cash_before",
            "position_cash_after",
            "position_total_value_after",
        ]

        # Write header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font

        # Write data
        for row_idx, row_data in enumerate(dividend_rows, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                value = row_data.get(col_name)
                if isinstance(value, datetime):
                    value = value.isoformat()
                elif isinstance(value, Decimal):
                    value = float(value)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

    def _create_summary_sheet(
        self, wb: Workbook, rows: List[Dict[str, Any]], tenant_id: str, portfolio_id: str
    ):
        """Create Summary sheet with KPIs, totals, drawdown, volatility."""
        ws = wb.create_sheet(title="Summary")

        if not rows:
            ws.cell(row=1, column=1, value="No data available")
            return

        # Calculate KPIs
        total_trades = len([r for r in rows if r.get("action") in ("BUY", "SELL")])
        buy_trades = len([r for r in rows if r.get("action") == "BUY"])
        sell_trades = len([r for r in rows if r.get("action") == "SELL"])

        total_commission = sum(
            float(r.get("commission_value") or 0.0) for r in rows if r.get("commission_value")
        )

        total_dividends = sum(
            float(r.get("dividend_net_value") or 0.0) for r in rows if r.get("dividend_net_value")
        )

        # Portfolio value progression
        portfolio_values = [
            (r.get("timestamp"), r.get("portfolio_total_value_before"))
            for r in rows
            if r.get("portfolio_total_value_before") is not None
        ]
        portfolio_values.sort(key=lambda x: x[0] or datetime.min)

        initial_value = portfolio_values[0][1] if portfolio_values else None
        final_value = portfolio_values[-1][1] if portfolio_values else None

        # Calculate drawdown
        max_value = max((v[1] for v in portfolio_values if v[1]), default=0.0)
        min_value_after_max = None
        if max_value > 0:
            max_idx = next(i for i, v in enumerate(portfolio_values) if v[1] == max_value)
            min_value_after_max = min(
                (v[1] for v in portfolio_values[max_idx:] if v[1]), default=max_value
            )
            max_drawdown_pct = (
                ((min_value_after_max - max_value) / max_value * 100)
                if min_value_after_max
                else 0.0
            )
        else:
            max_drawdown_pct = 0.0

        # Calculate returns
        total_return_pct = (
            ((final_value - initial_value) / initial_value * 100)
            if initial_value and final_value and initial_value > 0
            else 0.0
        )

        # Write summary
        summary_data = [
            ["Metric", "Value"],
            ["Total Evaluations", len(rows)],
            ["Total Trades", total_trades],
            ["Buy Trades", buy_trades],
            ["Sell Trades", sell_trades],
            ["Total Commission Paid", f"${total_commission:.2f}"],
            ["Total Dividends Received", f"${total_dividends:.2f}"],
            ["", ""],
            ["Portfolio Value", ""],
            ["Initial Value", f"${initial_value:.2f}" if initial_value else "N/A"],
            ["Final Value", f"${final_value:.2f}" if final_value else "N/A"],
            ["Total Return %", f"{total_return_pct:.2f}%"],
            ["Max Drawdown %", f"{max_drawdown_pct:.2f}%"],
        ]

        # Write data
        for row_idx, row_data in enumerate(summary_data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # Header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="366092", end_color="366092", fill_type="solid"
                    )
                    cell.font = Font(bold=True, color="FFFFFF")

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
