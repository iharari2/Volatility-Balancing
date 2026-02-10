# =========================
# backend/application/services/activity_report_export_service.py
# =========================
"""
Activity Report Export Service

Generates a unified Excel report consolidating:
- Trades from TradesRepo
- Events from EventsRepo
- Evaluation Timeline from EvaluationTimelineRepo

Creates a comprehensive activity report with multiple sheets:
- Summary: Report metadata and key metrics
- Activity Log: All events merged chronologically
- Trades: Actual executed trades
- Evaluation Timeline: Position evaluation data
- Dividends: Dividend events
"""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from domain.entities.event import Event
from domain.entities.trade import Trade
from domain.ports.events_repo import EventsRepo
from domain.ports.evaluation_timeline_repo import EvaluationTimelineRepo
from domain.ports.positions_repo import PositionsRepo
from domain.ports.trades_repo import TradesRepo


class ActivityReportExportService:
    """Service to generate unified activity reports in Excel format."""

    def __init__(
        self,
        trades_repo: TradesRepo,
        events_repo: EventsRepo,
        timeline_repo: EvaluationTimelineRepo,
        positions_repo: PositionsRepo,
    ):
        self.trades_repo = trades_repo
        self.events_repo = events_repo
        self.timeline_repo = timeline_repo
        self.positions_repo = positions_repo

    def export_activity_report(
        self,
        tenant_id: str,
        portfolio_id: str,
        position_id: Optional[str] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        mode: Optional[str] = None,
    ) -> bytes:
        """
        Generate comprehensive activity report Excel file.

        Args:
            tenant_id: Tenant identifier
            portfolio_id: Portfolio identifier
            position_id: Optional position filter
            start_date: Optional start date filter
            end_date: Optional end date filter
            mode: Optional mode filter (LIVE or SIMULATION)

        Returns:
            Excel file as bytes
        """
        # Fetch data from all sources
        trades = self._fetch_trades(tenant_id, portfolio_id, position_id, start_date, end_date)
        events = self._fetch_events(tenant_id, portfolio_id, position_id)
        timeline = self._fetch_timeline(
            tenant_id, portfolio_id, position_id, start_date, end_date, mode
        )

        # Merge into unified activity log
        activity_log = self._merge_to_activity_log(trades, events, timeline)

        # Filter dividends from timeline
        dividends = [t for t in timeline if t.get("dividend_applied", False)]

        # Create workbook
        return self._create_workbook(
            tenant_id=tenant_id,
            portfolio_id=portfolio_id,
            position_id=position_id,
            start_date=start_date,
            end_date=end_date,
            mode=mode,
            trades=trades,
            events=events,
            timeline=timeline,
            activity_log=activity_log,
            dividends=dividends,
        )

    def _fetch_trades(
        self,
        tenant_id: str,
        portfolio_id: str,
        position_id: Optional[str],
        start_date: Optional[datetime],
        end_date: Optional[datetime],
    ) -> List[Trade]:
        """Fetch trades from repository."""
        trades = self.trades_repo.list_by_portfolio(
            tenant_id=tenant_id,
            portfolio_id=portfolio_id,
            start_date=start_date,
            end_date=end_date,
        )
        if position_id:
            trades = [t for t in trades if t.position_id == position_id]
        return trades

    def _fetch_events(
        self,
        tenant_id: str,
        portfolio_id: str,
        position_id: Optional[str],
    ) -> List[Event]:
        """Fetch events from repository."""
        events_list: List[Event] = []
        if position_id:
            events_list = list(self.events_repo.list_for_position(position_id))
        else:
            # Get all positions for portfolio and fetch events for each
            positions = self.positions_repo.list_all(tenant_id=tenant_id, portfolio_id=portfolio_id)
            for pos in positions:
                pos_events = list(self.events_repo.list_for_position(pos.id))
                events_list.extend(pos_events)
        return events_list

    def _fetch_timeline(
        self,
        tenant_id: str,
        portfolio_id: str,
        position_id: Optional[str],
        start_date: Optional[datetime],
        end_date: Optional[datetime],
        mode: Optional[str],
    ) -> List[Dict[str, Any]]:
        """Fetch timeline data from repository."""
        if position_id:
            return self.timeline_repo.list_by_position(
                tenant_id=tenant_id,
                portfolio_id=portfolio_id,
                position_id=position_id,
                mode=mode,
                start_date=start_date,
                end_date=end_date,
            )
        else:
            return self.timeline_repo.list_by_portfolio(
                tenant_id=tenant_id,
                portfolio_id=portfolio_id,
                mode=mode,
                start_date=start_date,
                end_date=end_date,
            )

    def _merge_to_activity_log(
        self,
        trades: List[Trade],
        events: List[Event],
        timeline: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """Merge all sources into a unified chronological activity log."""
        activity_log: List[Dict[str, Any]] = []

        # Add trades
        for trade in trades:
            activity_log.append(
                {
                    "timestamp": trade.executed_at,
                    "event_type": "TRADE",
                    "position_id": trade.position_id,
                    "symbol": None,  # Will be enriched later if needed
                    "description": f"{trade.side} {trade.qty} @ ${trade.price:.2f}",
                    "details": {
                        "side": trade.side,
                        "qty": trade.qty,
                        "price": trade.price,
                        "commission": trade.commission,
                        "order_id": trade.order_id,
                    },
                    "related_trade_id": trade.id,
                }
            )

        # Add events
        for event in events:
            activity_log.append(
                {
                    "timestamp": event.ts,
                    "event_type": event.type,
                    "position_id": event.position_id,
                    "symbol": None,
                    "description": event.message,
                    "details": {"inputs": event.inputs, "outputs": event.outputs},
                    "related_trade_id": None,
                }
            )

        # Add timeline entries (excluding those already captured as trades)
        trade_ids = {t.id for t in trades}
        for entry in timeline:
            trade_id = entry.get("trade_id")
            if trade_id and trade_id in trade_ids:
                continue  # Skip, already have the trade

            # Determine event type from timeline action
            action = entry.get("action", "EVALUATION")
            if action in ("BUY", "SELL"):
                event_type = "TRIGGER"
            elif entry.get("dividend_applied"):
                event_type = "DIVIDEND"
            elif entry.get("guardrail_block_reason"):
                event_type = "GUARDRAIL"
            else:
                event_type = "EVALUATION"

            # Build description
            description_parts = []
            if entry.get("symbol"):
                description_parts.append(entry["symbol"])
            if action:
                description_parts.append(f"Action: {action}")
            if entry.get("action_reason"):
                description_parts.append(entry["action_reason"])

            activity_log.append(
                {
                    "timestamp": entry.get("timestamp"),
                    "event_type": event_type,
                    "position_id": entry.get("position_id"),
                    "symbol": entry.get("symbol"),
                    "description": " - ".join(description_parts) if description_parts else action,
                    "details": {
                        "effective_price": entry.get("effective_price"),
                        "anchor_price": entry.get("anchor_price"),
                        "pct_change_from_anchor": entry.get("pct_change_from_anchor"),
                        "trigger_fired": entry.get("trigger_fired"),
                        "trigger_reason": entry.get("trigger_reason"),
                    },
                    "related_trade_id": trade_id,
                }
            )

        # Sort chronologically
        activity_log.sort(key=lambda x: x.get("timestamp") or datetime.min)

        return activity_log

    def _create_workbook(
        self,
        tenant_id: str,
        portfolio_id: str,
        position_id: Optional[str],
        start_date: Optional[datetime],
        end_date: Optional[datetime],
        mode: Optional[str],
        trades: List[Trade],
        events: List[Event],
        timeline: List[Dict[str, Any]],
        activity_log: List[Dict[str, Any]],
        dividends: List[Dict[str, Any]],
    ) -> bytes:
        """Create Excel workbook with all sheets."""
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Sheet 1: Summary
        self._create_summary_sheet(
            wb,
            tenant_id=tenant_id,
            portfolio_id=portfolio_id,
            position_id=position_id,
            start_date=start_date,
            end_date=end_date,
            mode=mode,
            trades=trades,
            dividends=dividends,
            timeline=timeline,
        )

        # Sheet 2: Activity Log
        self._create_activity_log_sheet(wb, activity_log)

        # Sheet 3: Trades
        self._create_trades_sheet(wb, trades)

        # Sheet 4: Evaluation Timeline
        self._create_timeline_sheet(wb, timeline)

        # Sheet 5: Dividends
        self._create_dividends_sheet(wb, dividends)

        # Save to bytes
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _create_summary_sheet(
        self,
        wb: Workbook,
        tenant_id: str,
        portfolio_id: str,
        position_id: Optional[str],
        start_date: Optional[datetime],
        end_date: Optional[datetime],
        mode: Optional[str],
        trades: List[Trade],
        dividends: List[Dict[str, Any]],
        timeline: List[Dict[str, Any]],
    ) -> None:
        """Create Summary sheet with report metadata and key metrics."""
        ws = wb.create_sheet(title="Summary")

        # Calculate metrics
        total_trades = len(trades)
        buy_trades = len([t for t in trades if t.side == "BUY"])
        sell_trades = len([t for t in trades if t.side == "SELL"])
        total_commission = sum(t.commission for t in trades)
        total_dividends = sum(
            float(d.get("dividend_net_value") or 0.0) for d in dividends
        )

        # Calculate P&L from timeline if available
        pnl = 0.0
        if timeline:
            first_value = None
            last_value = None
            sorted_timeline = sorted(timeline, key=lambda x: x.get("timestamp") or datetime.min)
            for entry in sorted_timeline:
                val = entry.get("position_total_value_before")
                if val is not None:
                    if first_value is None:
                        first_value = float(val)
                    last_value = float(val)
            if first_value and last_value:
                pnl = last_value - first_value

        # Summary data
        summary_data = [
            ["Activity Report", ""],
            ["", ""],
            ["Report Parameters", ""],
            ["Tenant ID", tenant_id],
            ["Portfolio ID", portfolio_id],
            ["Position ID", position_id or "All positions"],
            ["Start Date", start_date.isoformat() if start_date else "Not specified"],
            ["End Date", end_date.isoformat() if end_date else "Not specified"],
            ["Mode", mode or "All modes"],
            ["Generated At", datetime.now().isoformat()],
            ["", ""],
            ["Key Metrics", ""],
            ["Total Trades", total_trades],
            ["Buy Trades", buy_trades],
            ["Sell Trades", sell_trades],
            ["Total Commission", f"${total_commission:.2f}"],
            ["Total Dividends", f"${total_dividends:.2f}"],
            ["Estimated P&L", f"${pnl:.2f}"],
        ]

        # Write data with styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        section_font = Font(bold=True, size=12)

        for row_idx, row_data in enumerate(summary_data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Style section headers
                if col_idx == 1 and value in ("Activity Report", "Report Parameters", "Key Metrics"):
                    cell.font = section_font

        # Set column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30

    def _create_activity_log_sheet(self, wb: Workbook, activity_log: List[Dict[str, Any]]) -> None:
        """Create Activity Log sheet with all events merged chronologically."""
        ws = wb.create_sheet(title="Activity Log")

        columns = [
            "timestamp",
            "event_type",
            "position_id",
            "symbol",
            "description",
            "details",
            "related_trade_id",
        ]

        # Write header
        self._write_header(ws, columns)

        # Write data
        for row_idx, row_data in enumerate(activity_log, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                value = row_data.get(col_name)
                if isinstance(value, datetime):
                    value = value.isoformat()
                elif isinstance(value, dict):
                    import json
                    value = json.dumps(value, default=str)
                elif isinstance(value, Decimal):
                    value = float(value)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        self._auto_adjust_columns(ws, columns)

    def _create_trades_sheet(self, wb: Workbook, trades: List[Trade]) -> None:
        """Create Trades sheet with actual executed trades."""
        ws = wb.create_sheet(title="Trades")

        columns = [
            "executed_at",
            "position_id",
            "side",
            "qty",
            "price",
            "notional",
            "commission",
            "order_id",
            "trade_id",
        ]

        # Write header
        self._write_header(ws, columns)

        # Write data
        for row_idx, trade in enumerate(trades, start=2):
            notional = trade.qty * trade.price
            row_values = [
                trade.executed_at.isoformat() if trade.executed_at else None,
                trade.position_id,
                trade.side,
                trade.qty,
                trade.price,
                notional,
                trade.commission,
                trade.order_id,
                trade.id,
            ]
            for col_idx, value in enumerate(row_values, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        self._auto_adjust_columns(ws, columns)

    def _create_timeline_sheet(self, wb: Workbook, timeline: List[Dict[str, Any]]) -> None:
        """Create Evaluation Timeline sheet (non-verbose view)."""
        ws = wb.create_sheet(title="Evaluation Timeline")

        # Non-verbose columns (key information)
        columns = [
            "timestamp",
            "mode",
            "symbol",
            "position_id",
            "effective_price",
            "position_qty_before",
            "position_cash_before",
            "position_total_value_before",
            "stock_pct",
            "anchor_price",
            "pct_change_from_anchor",
            "trigger_fired",
            "trigger_direction",
            "action",
            "action_reason",
            "trade_intent_qty",
            "execution_price",
            "execution_qty",
            "position_qty_after",
            "position_cash_after",
            "position_total_value_after",
        ]

        # Write header
        self._write_header(ws, columns)

        # Sort by timestamp
        sorted_timeline = sorted(timeline, key=lambda x: x.get("timestamp") or datetime.min)

        # Write data
        for row_idx, row_data in enumerate(sorted_timeline, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                value = row_data.get(col_name)
                if isinstance(value, datetime):
                    value = value.isoformat()
                elif isinstance(value, Decimal):
                    value = float(value)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        self._auto_adjust_columns(ws, columns)

    def _create_dividends_sheet(self, wb: Workbook, dividends: List[Dict[str, Any]]) -> None:
        """Create Dividends sheet."""
        ws = wb.create_sheet(title="Dividends")

        columns = [
            "timestamp",
            "symbol",
            "position_id",
            "dividend_ex_date",
            "dividend_pay_date",
            "dividend_rate",
            "dividend_gross_value",
            "dividend_tax",
            "dividend_net_value",
            "position_qty_before",
            "position_cash_before",
            "position_cash_after",
        ]

        # Write header
        self._write_header(ws, columns)

        # Sort by timestamp
        sorted_dividends = sorted(dividends, key=lambda x: x.get("timestamp") or datetime.min)

        # Write data
        for row_idx, row_data in enumerate(sorted_dividends, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                value = row_data.get(col_name)
                if isinstance(value, datetime):
                    value = value.isoformat()
                elif isinstance(value, Decimal):
                    value = float(value)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        self._auto_adjust_columns(ws, columns)

    def _write_header(self, ws, columns: List[str]) -> None:
        """Write styled header row."""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _auto_adjust_columns(self, ws, columns: List[str]) -> None:
        """Auto-adjust column widths."""
        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
