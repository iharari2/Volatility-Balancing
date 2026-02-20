# =========================
# backend/application/services/excel_template_service.py
# =========================

import io
import json
from datetime import datetime, timezone
from typing import Dict, List, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

from domain.entities.optimization_result import OptimizationResults
from domain.entities.simulation_result import SimulationResult
from application.services.verbose_timeline_service import VerboseTimelineService


class ExcelTemplateService:
    """Service for creating professional Excel templates with advanced formatting."""

    def __init__(self):
        self.styles = {}

    def _safe_auto_fit_columns(self, ws):
        """Safely auto-fit columns, handling merged cells properly"""
        for column in ws.columns:
            max_length = 0
            column_letter = None

            # Find the first non-merged cell to get column letter
            for cell in column:
                if hasattr(cell, "column_letter") and column_letter is None:
                    column_letter = cell.column_letter
                try:
                    if hasattr(cell, "value") and cell.value is not None:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except (TypeError, ValueError, AttributeError):
                    pass

            # Only adjust width if we found a valid column letter
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

    def _old_init_method(self):
        """Initialize the Excel template service."""
        self.workbook = None
        self.styles = {}

    def create_optimization_report(
        self, results: OptimizationResults, config_name: str = "Optimization Results"
    ) -> bytes:
        """Create a comprehensive optimization report with professional formatting."""
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)

        # Define styles
        self._define_styles()

        # Create sheets
        self._create_optimization_executive_summary(results, config_name)
        self._create_optimization_detailed_results(results)
        self._create_parameter_sensitivity_analysis(results)
        self._create_performance_metrics_analysis(results)
        self._create_optimization_heatmap_data(results)
        self._create_optimization_recommendations(results)

        # Save to bytes
        excel_buffer = io.BytesIO()
        self.workbook.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()

    def create_simulation_report(
        self, simulation_result: SimulationResult, ticker: str = "UNKNOWN"
    ) -> bytes:
        """Create a comprehensive simulation report with professional formatting."""
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)

        # Define styles
        self._define_styles()

        # Create sheets
        self._create_simulation_executive_summary(simulation_result, ticker)
        self._create_verbose_timeline(simulation_result, ticker)
        self._create_comprehensive_time_series(simulation_result)
        self._create_simulation_performance_analysis(simulation_result)
        self._create_detailed_trade_log(simulation_result)
        self._create_dividend_events(simulation_result)
        self._create_trigger_analysis(simulation_result)
        self._create_daily_returns(simulation_result)
        self._create_price_data(simulation_result)
        self._create_risk_analysis(simulation_result)
        self._create_market_analysis(simulation_result)

        # Save to bytes
        excel_buffer = io.BytesIO()
        self.workbook.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()

    def create_trading_audit_report(
        self,
        positions_data: List[Dict[str, Any]],
        trades_data: List[Dict[str, Any]],
        orders_data: List[Dict[str, Any]],
        title: str = "Trading Audit Report",
    ) -> bytes:
        """Create a comprehensive trading audit report."""
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)

        # Define styles
        self._define_styles()

        # Create sheets
        self._create_trading_executive_summary(positions_data, trades_data, orders_data, title)
        self._create_positions_analysis(positions_data)
        self._create_trades_analysis(trades_data)
        self._create_orders_analysis(orders_data)
        self._create_compliance_analysis(positions_data, trades_data, orders_data)
        self._create_performance_attribution(positions_data, trades_data)

        # Save to bytes
        excel_buffer = io.BytesIO()
        self.workbook.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()

    def _define_styles(self):
        """Define custom styles for the workbook."""
        # Header style
        header_style = NamedStyle(name="header")
        header_style.font = Font(bold=True, color="FFFFFF", size=12)
        header_style.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_style.alignment = Alignment(horizontal="center", vertical="center")
        header_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self.styles["header"] = header_style

        # Subheader style
        subheader_style = NamedStyle(name="subheader")
        subheader_style.font = Font(bold=True, color="2F5597", size=11)
        subheader_style.fill = PatternFill(
            start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
        )
        subheader_style.alignment = Alignment(horizontal="left", vertical="center")
        self.styles["subheader"] = subheader_style

        # Data style
        data_style = NamedStyle(name="data")
        data_style.font = Font(size=10)
        data_style.alignment = Alignment(horizontal="left", vertical="center")
        data_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self.styles["data"] = data_style

        # Number style
        number_style = NamedStyle(name="number")
        number_style.font = Font(size=10)
        number_style.alignment = Alignment(horizontal="right", vertical="center")
        number_style.number_format = "#,##0.00"
        number_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self.styles["number"] = number_style

        # Percentage style
        percentage_style = NamedStyle(name="percentage")
        percentage_style.font = Font(size=10)
        percentage_style.alignment = Alignment(horizontal="right", vertical="center")
        percentage_style.number_format = "0.00%"
        percentage_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self.styles["percentage"] = percentage_style

    def _create_optimization_executive_summary(
        self, results: OptimizationResults, config_name: str
    ):
        """Create executive summary sheet for optimization results."""
        ws = self.workbook.create_sheet("Executive Summary", 0)

        # Title
        ws["A1"] = f"{config_name} - Executive Summary"
        ws["A1"].font = Font(size=18, bold=True, color="2F5597")
        ws.merge_cells("A1:H1")

        # Generation info
        ws["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}"
        ws["A2"].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells("A2:H2")

        # Key metrics section
        ws["A4"] = "Key Performance Metrics"
        ws["A4"].font = Font(size=14, bold=True, color="2F5597")
        ws.merge_cells("A4:H4")

        # Summary statistics
        row = 6
        ws[f"A{row}"] = "Total Parameter Combinations:"
        ws[f"B{row}"] = len(results.results)
        ws[f"C{row}"] = "Completed:"
        ws[f"D{row}"] = len(results.get_completed_results())
        ws[f"E{row}"] = "Failed:"
        ws[f"F{row}"] = len(results.get_failed_results())
        ws[f"G{row}"] = "Success Rate:"
        ws[f"H{row}"] = (
            f"{len(results.get_completed_results()) / len(results.results) * 100:.1f}%"
            if results.results
            else "0%"
        )

        # Best result summary
        if results.best_result:
            row += 3
            ws[f"A{row}"] = "Best Performing Configuration"
            ws[f"A{row}"].font = Font(size=12, bold=True, color="2F5597")
            ws.merge_cells(f"A{row}:H{row}")

            row += 1
            ws[f"A{row}"] = "Combination ID:"
            ws[f"B{row}"] = results.best_result.parameter_combination.combination_id
            ws[f"C{row}"] = "Status:"
            ws[f"D{row}"] = results.best_result.status.value.title()

            # Parameters
            row += 2
            ws[f"A{row}"] = "Parameters:"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

            for param, value in results.best_result.parameter_combination.parameters.items():
                ws[f"A{row}"] = f"  {param}:"
                ws[f"B{row}"] = str(value)
                row += 1

            # Metrics
            row += 1
            ws[f"A{row}"] = "Performance Metrics:"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

            for metric, value in results.best_result.metrics.items():
                ws[f"A{row}"] = f"  {metric.value}:"
                ws[f"B{row}"] = f"{value:.6f}"
                row += 1

        # Apply styles
        self._apply_styles_to_range(ws, "A1:H50")

    def _create_optimization_detailed_results(self, results: OptimizationResults):
        """Create detailed results sheet with advanced formatting."""
        ws = self.workbook.create_sheet("Detailed Results")

        if not results.results:
            ws["A1"] = "No results available"
            return

        # Prepare data for DataFrame
        data = []
        for result in results.results:
            row = {
                "ID": str(result.id),
                "Combination ID": result.parameter_combination.combination_id,
                "Status": result.status.value.title(),
                "Created At": result.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                "Completed At": (
                    result.completed_at.strftime("%Y-%m-%d %H:%M:%S") if result.completed_at else ""
                ),
                "Execution Time (s)": result.execution_time_seconds or "",
                "Error Message": result.error_message or "",
            }

            # Add parameters
            for param, value in result.parameter_combination.parameters.items():
                row[f"Param_{param}"] = value

            # Add metrics
            for metric, value in result.metrics.items():
                row[f"Metric_{metric.value}"] = value

            data.append(row)

        # Create DataFrame and write to Excel
        df = pd.DataFrame(data)

        # Write headers with styling
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.style = self.styles["header"]

        # Write data with conditional formatting
        for row_num, row_data in enumerate(df.itertuples(index=False), 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)

                # Apply appropriate style based on column type
                if (
                    "Metric_" in df.columns[col_num - 1]
                    or "Execution Time" in df.columns[col_num - 1]
                ):
                    cell.style = self.styles["number"]
                elif "Status" in df.columns[col_num - 1]:
                    cell.style = self.styles["data"]
                else:
                    cell.style = self.styles["data"]

        # Auto-adjust column widths safely
        self._safe_auto_fit_columns(ws)

        # Add conditional formatting for metrics
        self._add_conditional_formatting(ws, df)

    def _create_parameter_sensitivity_analysis(self, results: OptimizationResults):
        """Create parameter sensitivity analysis sheet."""
        ws = self.workbook.create_sheet("Parameter Sensitivity")

        if not results.results:
            ws["A1"] = "No results available"
            return

        # Get all unique parameters
        all_params = set()
        for result in results.results:
            all_params.update(result.parameter_combination.parameters.keys())

        # Create parameter statistics
        ws["A1"] = "Parameter Sensitivity Analysis"
        ws["A1"].font = Font(size=16, bold=True, color="2F5597")
        ws.merge_cells("A1:H1")

        row = 3
        for param in sorted(all_params):
            ws[f"A{row}"] = f"Parameter: {param}"
            ws[f"A{row}"].font = Font(size=12, bold=True, color="2F5597")
            ws.merge_cells(f"A{row}:H{row}")
            row += 1

            # Get all values for this parameter
            values = []
            for result in results.results:
                if param in result.parameter_combination.parameters:
                    values.append(result.parameter_combination.parameters[param])

            if values:
                # Create statistics table
                stats_data = [
                    ["Statistic", "Value"],
                    ["Count", len(values)],
                    ["Unique Values", len(set(values))],
                ]

                if all(isinstance(v, (int, float)) for v in values):
                    stats_data.extend(
                        [
                            ["Minimum", min(values)],
                            ["Maximum", max(values)],
                            ["Mean", sum(values) / len(values)],
                            ["Range", max(values) - min(values)],
                        ]
                    )

                # Write statistics table
                for i, (stat, val) in enumerate(stats_data):
                    ws[f"B{row}"] = stat
                    ws[f"C{row}"] = val
                    if i == 0:  # Header row
                        ws[f"B{row}"].style = self.styles["header"]
                        ws[f"C{row}"].style = self.styles["header"]
                    else:
                        ws[f"B{row}"].style = self.styles["data"]
                        ws[f"C{row}"].style = self.styles["number"]
                    row += 1

                row += 2

    def _create_performance_metrics_analysis(self, results: OptimizationResults):
        """Create performance metrics analysis sheet."""
        ws = self.workbook.create_sheet("Performance Metrics")

        if not results.results:
            ws["A1"] = "No results available"
            return

        # Get all unique metrics
        all_metrics = set()
        for result in results.results:
            all_metrics.update(result.metrics.keys())

        # Create metrics statistics
        ws["A1"] = "Performance Metrics Analysis"
        ws["A1"].font = Font(size=16, bold=True, color="2F5597")
        ws.merge_cells("A1:H1")

        row = 3
        for metric in sorted(all_metrics, key=lambda x: x.value):
            ws[f"A{row}"] = f"Metric: {metric.value}"
            ws[f"A{row}"].font = Font(size=12, bold=True, color="2F5597")
            ws.merge_cells(f"A{row}:H{row}")
            row += 1

            # Get all values for this metric
            values = []
            for result in results.results:
                if metric in result.metrics:
                    values.append(result.metrics[metric])

            if values:
                # Create statistics table
                stats_data = [
                    ["Statistic", "Value"],
                    ["Count", len(values)],
                    ["Minimum", f"{min(values):.6f}"],
                    ["Maximum", f"{max(values):.6f}"],
                    ["Mean", f"{sum(values) / len(values):.6f}"],
                    ["Range", f"{max(values) - min(values):.6f}"],
                ]

                # Write statistics table
                for i, (stat, val) in enumerate(stats_data):
                    ws[f"B{row}"] = stat
                    ws[f"C{row}"] = val
                    if i == 0:  # Header row
                        ws[f"B{row}"].style = self.styles["header"]
                        ws[f"C{row}"].style = self.styles["header"]
                    else:
                        ws[f"B{row}"].style = self.styles["data"]
                        ws[f"C{row}"].style = self.styles["number"]
                    row += 1

                row += 2

    def _create_optimization_heatmap_data(self, results: OptimizationResults):
        """Create heatmap data sheet for visualization."""
        ws = self.workbook.create_sheet("Heatmap Data")

        if not results.results:
            ws["A1"] = "No results available"
            return

        # Get all unique parameters
        all_params = list(set())
        for result in results.results:
            all_params.extend(result.parameter_combination.parameters.keys())
        all_params = list(set(all_params))

        # Get all unique metrics
        all_metrics = list(set())
        for result in results.results:
            all_metrics.extend(result.metrics.keys())
        all_metrics = list(set(all_metrics))

        # Create parameter combinations matrix
        ws["A1"] = "Parameter Combinations Matrix"
        ws["A1"].font = Font(size=14, bold=True, color="2F5597")

        # Create headers
        headers = ["Combination ID"] + all_params + [metric.value for metric in all_metrics]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.style = self.styles["header"]

        # Fill data
        row = 3
        for result in results.results:
            ws.cell(row=row, column=1, value=result.parameter_combination.combination_id)

            # Add parameter values
            for col_num, param in enumerate(all_params, 2):
                value = result.parameter_combination.parameters.get(param, "")
                ws.cell(row=row, column=col_num, value=value)

            # Add metric values
            for col_num, metric in enumerate(all_metrics, 2 + len(all_params)):
                value = result.metrics.get(metric, "")
                ws.cell(row=row, column=col_num, value=value)

            row += 1

    def _create_optimization_recommendations(self, results: OptimizationResults):
        """Create recommendations sheet."""
        ws = self.workbook.create_sheet("Recommendations")

        ws["A1"] = "Optimization Recommendations"
        ws["A1"].font = Font(size=16, bold=True, color="2F5597")
        ws.merge_cells("A1:H1")

        # Best result recommendations
        if results.best_result:
            ws["A3"] = "Recommended Configuration"
            ws["A3"].font = Font(size=12, bold=True, color="2F5597")
            ws.merge_cells("A3:H3")

            ws["A4"] = (
                f"Use Combination ID: {results.best_result.parameter_combination.combination_id}"
            )
            ws["A5"] = "Parameters:"

            row = 6
            for param, value in results.best_result.parameter_combination.parameters.items():
                ws[f"A{row}"] = f"  {param}: {value}"
                row += 1

            ws[f"A{row+1}"] = "Expected Performance:"
            row += 2
            for metric, value in results.best_result.metrics.items():
                ws[f"A{row}"] = f"  {metric.value}: {value:.6f}"
                row += 1

    def _create_simulation_executive_summary(
        self, simulation_result: SimulationResult, ticker: str
    ):
        """Create executive summary for simulation results."""
        ws = self.workbook.create_sheet("Executive Summary", 0)

        # Title
        ws["A1"] = f"Simulation Results - {ticker}"
        ws["A1"].font = Font(size=18, bold=True, color="2F5597")
        ws.merge_cells("A1:H1")

        # Generation info - use fully qualified name to avoid shadowing issues
        from datetime import datetime as dt_module, timezone as tz_module

        ws["A2"] = f"Generated: {dt_module.now(tz_module.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}"
        ws["A2"].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells("A2:H2")

        # Simulation details
        ws["A4"] = "Simulation Details"
        ws["A4"].font = Font(size=14, bold=True, color="2F5597")
        ws.merge_cells("A4:H4")

        ws["A5"] = "Ticker:"
        ws["B5"] = ticker
        ws["A6"] = "Start Date:"
        ws["B6"] = simulation_result.start_date
        ws["A7"] = "End Date:"
        ws["B7"] = simulation_result.end_date
        ws["A8"] = "Total Trading Days:"
        # Calculate trading days from date range
        from datetime import datetime

        start_date = datetime.strptime(simulation_result.start_date, "%Y-%m-%d")
        end_date = datetime.strptime(simulation_result.end_date, "%Y-%m-%d")
        trading_days = (end_date - start_date).days
        ws["B8"] = trading_days
        ws["A9"] = "Initial Cash:"
        # Get initial cash from raw_data or use default
        initial_cash = (
            simulation_result.raw_data.get("initial_cash", 100000)
            if simulation_result.raw_data
            else 100000
        )
        ws["B9"] = f"${initial_cash:,.2f}"

        # Algorithm Performance
        ws["A11"] = "Algorithm Performance"
        ws["A11"].font = Font(size=14, bold=True, color="2F5597")
        ws.merge_cells("A11:H11")

        # Get data from metrics or raw_data
        algorithm_trades = (
            simulation_result.raw_data.get("algorithm_trades", 0)
            if simulation_result.raw_data
            else 0
        )
        algorithm_pnl = simulation_result.metrics.get("algorithm_pnl", 0)
        algorithm_return_pct = simulation_result.metrics.get("total_return", 0) * 100
        algorithm_volatility = simulation_result.metrics.get("volatility", 0) * 100
        algorithm_sharpe_ratio = simulation_result.metrics.get("sharpe_ratio", 0)
        algorithm_max_drawdown = simulation_result.metrics.get("max_drawdown", 0) * 100

        ws["A12"] = "Total Trades:"
        ws["B12"] = algorithm_trades
        ws["A13"] = "Algorithm P&L:"
        ws["B13"] = f"${algorithm_pnl:,.2f}"
        ws["A14"] = "Algorithm Return:"
        ws["B14"] = f"{algorithm_return_pct:.2f}%"
        ws["A15"] = "Algorithm Volatility:"
        ws["B15"] = f"{algorithm_volatility:.2f}%"
        ws["A16"] = "Algorithm Sharpe Ratio:"
        ws["B16"] = f"{algorithm_sharpe_ratio:.2f}"
        ws["A17"] = "Algorithm Max Drawdown:"
        ws["B17"] = f"{algorithm_max_drawdown:.2f}%"

        # Buy & Hold Performance
        ws["A19"] = "Buy & Hold Performance"
        ws["A19"].font = Font(size=14, bold=True, color="2F5597")
        ws.merge_cells("A19:H19")

        # Get buy & hold data from metrics or use defaults
        buy_hold_pnl = simulation_result.metrics.get("buy_hold_pnl", 0)
        buy_hold_return_pct = simulation_result.metrics.get("buy_hold_return_pct", 0)
        buy_hold_volatility = simulation_result.metrics.get("buy_hold_volatility", 0)
        buy_hold_sharpe_ratio = simulation_result.metrics.get("buy_hold_sharpe_ratio", 0)
        buy_hold_max_drawdown = simulation_result.metrics.get("buy_hold_max_drawdown", 0)

        ws["A20"] = "Buy & Hold P&L:"
        ws["B20"] = f"${buy_hold_pnl:,.2f}"
        ws["A21"] = "Buy & Hold Return:"
        ws["B21"] = f"{buy_hold_return_pct:.2f}%"
        ws["A22"] = "Buy & Hold Volatility:"
        ws["B22"] = f"{buy_hold_volatility:.2f}%"
        ws["A23"] = "Buy & Hold Sharpe Ratio:"
        ws["B23"] = f"{buy_hold_sharpe_ratio:.2f}"
        ws["A24"] = "Buy & Hold Max Drawdown:"
        ws["B24"] = f"{buy_hold_max_drawdown:.2f}%"

        # Comparison
        ws["A26"] = "Performance Comparison"
        ws["A26"].font = Font(size=14, bold=True, color="2F5597")
        ws.merge_cells("A26:H26")

        # Calculate comparison metrics
        excess_return = algorithm_return_pct - buy_hold_return_pct
        alpha = simulation_result.metrics.get("alpha", 0)
        beta = simulation_result.metrics.get("beta", 1)
        information_ratio = simulation_result.metrics.get("information_ratio", 0)

        ws["A27"] = "Excess Return:"
        ws["B27"] = f"{excess_return:.2f}%"
        ws["A28"] = "Alpha:"
        ws["B28"] = f"{alpha:.2f}"
        ws["A29"] = "Beta:"
        ws["B29"] = f"{beta:.2f}"
        ws["A30"] = "Information Ratio:"
        ws["B30"] = f"{information_ratio:.2f}"

        # Dividends
        total_dividends_received = (
            simulation_result.raw_data.get("total_dividends_received", 0)
            if simulation_result.raw_data
            else 0
        )
        if total_dividends_received > 0:
            ws["A32"] = "Dividend Information"
            ws["A32"].font = Font(size=14, bold=True, color="2F5597")
            ws.merge_cells("A32:H32")
            ws["A33"] = "Total Dividends Received:"
            ws["B33"] = f"${total_dividends_received:,.2f}"

    def _create_simulation_performance_analysis(self, simulation_result: SimulationResult):
        """Create performance analysis sheet."""
        ws = self.workbook.create_sheet("Performance Analysis")

        ws["A1"] = "Performance Analysis"
        ws["A1"].font = Font(size=16, bold=True, color="2F5597")
        ws.merge_cells("A1:H1")

        # Add performance analysis content here
        ws["A3"] = "Performance analysis will be implemented based on available simulation data."

    def _create_trade_analysis(self, simulation_result: SimulationResult):
        """Create trade analysis sheet."""
        ws = self.workbook.create_sheet("Trade Analysis")

        ws["A1"] = "Trade Analysis"
        ws["A1"].font = Font(size=16, bold=True, color="2F5597")
        ws.merge_cells("A1:H1")

        # Add trade analysis content here
        ws["A3"] = "Trade analysis will be implemented based on available trade data."

    def _create_risk_analysis(self, simulation_result: SimulationResult):
        """Create risk analysis sheet."""
        ws = self.workbook.create_sheet("Risk Analysis")

        ws["A1"] = "Risk Analysis"
        ws["A1"].font = Font(size=16, bold=True, color="2F5597")
        ws.merge_cells("A1:H1")

        # Add risk analysis content here
        ws["A3"] = "Risk analysis will be implemented based on available simulation data."

    def _create_market_analysis(self, simulation_result: SimulationResult):
        """Create market analysis sheet."""
        ws = self.workbook.create_sheet("Market Analysis")

        ws["A1"] = "Market Analysis"
        ws["A1"].font = Font(size=16, bold=True, color="2F5597")
        ws.merge_cells("A1:H1")

        # Add market analysis content here
        ws["A3"] = "Market analysis will be implemented based on available market data."

    def _create_trading_executive_summary(self, positions_data, trades_data, orders_data, title):
        """Create trading executive summary."""
        ws = self.workbook.create_sheet("Executive Summary", 0)

        ws["A1"] = title
        ws["A1"].font = Font(size=18, bold=True, color="2F5597")
        ws.merge_cells("A1:H1")

        ws["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}"
        ws["A2"].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells("A2:H2")

        # Summary statistics
        ws["A4"] = "Summary Statistics"
        ws["A4"].font = Font(size=14, bold=True, color="2F5597")
        ws.merge_cells("A4:H4")

        ws["A5"] = "Total Positions:"
        ws["B5"] = len(positions_data)
        ws["A6"] = "Total Trades:"
        ws["B6"] = len(trades_data)
        ws["A7"] = "Total Orders:"
        ws["B7"] = len(orders_data)

    def _create_positions_analysis(self, positions_data):
        """Create positions analysis sheet."""
        ws = self.workbook.create_sheet("Positions Analysis")

        ws["A1"] = "Positions Analysis"
        ws["A1"].font = Font(size=16, bold=True, color="2F5597")
        ws.merge_cells("A1:H1")

        # Add positions analysis content here
        ws["A3"] = "Positions analysis will be implemented based on available positions data."

    def _create_trades_analysis(self, trades_data):
        """Create trades analysis sheet."""
        ws = self.workbook.create_sheet("Trades Analysis")

        ws["A1"] = "Trades Analysis"
        ws["A1"].font = Font(size=16, bold=True, color="2F5597")
        ws.merge_cells("A1:H1")

        # Add trades analysis content here
        ws["A3"] = "Trades analysis will be implemented based on available trades data."

    def _create_orders_analysis(self, orders_data):
        """Create orders analysis sheet."""
        ws = self.workbook.create_sheet("Orders Analysis")

        ws["A1"] = "Orders Analysis"
        ws["A1"].font = Font(size=16, bold=True, color="2F5597")
        ws.merge_cells("A1:H1")

        # Add orders analysis content here
        ws["A3"] = "Orders analysis will be implemented based on available orders data."

    def _create_compliance_analysis(self, positions_data, trades_data, orders_data):
        """Create compliance analysis sheet."""
        ws = self.workbook.create_sheet("Compliance Analysis")

        ws["A1"] = "Compliance Analysis"
        ws["A1"].font = Font(size=16, bold=True, color="2F5597")
        ws.merge_cells("A1:H1")

        # Add compliance analysis content here
        ws["A3"] = "Compliance analysis will be implemented based on available trading data."

    def _create_performance_attribution(self, positions_data, trades_data):
        """Create performance attribution sheet."""
        ws = self.workbook.create_sheet("Performance Attribution")

        ws["A1"] = "Performance Attribution"
        ws["A1"].font = Font(size=16, bold=True, color="2F5597")
        ws.merge_cells("A1:H1")

        # Add performance attribution content here
        ws["A3"] = "Performance attribution will be implemented based on available trading data."

    def _apply_styles_to_range(self, ws, range_str):
        """Apply styles to a range of cells."""
        # This would apply styles to a range of cells
        pass

    def _add_conditional_formatting(self, ws, df):
        """Add conditional formatting to the worksheet."""
        # Add color scales for numeric columns
        for col_num, column in enumerate(df.columns, 1):
            if "Metric_" in column:
                # Apply color scale to metric columns
                col_letter = get_column_letter(col_num)
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{len(df)+1}",
                    ColorScaleRule(
                        start_type="min",
                        start_color="FF6B6B",
                        mid_type="percentile",
                        mid_value=50,
                        mid_color="FFE66D",
                        end_type="max",
                        end_color="4ECDC4",
                    ),
                )

    def _create_detailed_trade_log(self, simulation_result: SimulationResult):
        """Create detailed trade log sheet with all transactions."""
        ws = self.workbook.create_sheet("Detailed Trade Log")

        ws["A1"] = "Detailed Trade Log"
        ws["A1"].font = Font(size=16, bold=True, color="2F5597")
        ws.merge_cells("A1:H1")

        ws["A2"] = (
            f"Ticker: {simulation_result.ticker} | Period: {simulation_result.start_date} to {simulation_result.end_date}"
        )
        ws["A2"].font = Font(size=12, italic=True, color="666666")
        ws.merge_cells("A2:H2")

        trade_log = (
            simulation_result.raw_data.get("trade_log", []) if simulation_result.raw_data else []
        )
        if not trade_log:
            ws["A4"] = "No trade data available"
            return

        # Headers
        headers = [
            "Timestamp",
            "Side",
            "Quantity",
            "Price",
            "Commission",
            "Cash After",
            "Shares After",
            "Notional Value",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.style = self.styles["header"]

        # Trade data
        for row, trade in enumerate(trade_log, 5):
            ws.cell(row=row, column=1, value=trade.get("timestamp", ""))
            ws.cell(row=row, column=2, value=trade.get("side", ""))
            ws.cell(row=row, column=3, value=trade.get("qty", 0))
            ws.cell(row=row, column=4, value=trade.get("price", 0))
            ws.cell(row=row, column=5, value=trade.get("commission", 0))
            ws.cell(row=row, column=6, value=trade.get("cash_after", 0))
            ws.cell(row=row, column=7, value=trade.get("shares_after", 0))

            # Calculate notional value
            notional = trade.get("qty", 0) * trade.get("price", 0)
            ws.cell(row=row, column=8, value=notional)

            # Apply number formatting
            for col in [3, 4, 5, 6, 7, 8]:
                ws.cell(row=row, column=col).style = self.styles["number"]

        # Auto-adjust column widths safely
        self._safe_auto_fit_columns(ws)

    def _create_dividend_events(self, simulation_result: SimulationResult):
        """Create dividend events sheet with all dividend transactions."""
        ws = self.workbook.create_sheet("Dividend Events")

        ws["A1"] = "Dividend Events"
        ws["A1"].font = Font(size=16, bold=True, color="2F5597")
        ws.merge_cells("A1:H1")

        ws["A2"] = (
            f"Ticker: {simulation_result.ticker} | Total Dividends: ${getattr(simulation_result, 'total_dividends_received', 0):,.2f}"
        )
        ws["A2"].font = Font(size=12, italic=True, color="666666")
        ws.merge_cells("A2:H2")

        dividend_events = (
            simulation_result.raw_data.get("dividend_events", [])
            if simulation_result.raw_data
            else []
        )
        if not dividend_events:
            ws["A4"] = "No dividend events during simulation period"
            return

        # Headers
        headers = [
            "Date",
            "Ex-Date",
            "Pay Date",
            "DPS",
            "Shares",
            "Gross Amount",
            "Net Amount",
            "Withholding Tax",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.style = self.styles["header"]

        # Dividend data
        for row, dividend in enumerate(dividend_events, 5):
            ws.cell(row=row, column=1, value=dividend.get("date", ""))
            ws.cell(row=row, column=2, value=dividend.get("ex_date", ""))
            ws.cell(row=row, column=3, value=dividend.get("pay_date", ""))
            ws.cell(row=row, column=4, value=dividend.get("dps", 0))
            ws.cell(row=row, column=5, value=dividend.get("shares", 0))
            ws.cell(row=row, column=6, value=dividend.get("gross_amount", 0))
            ws.cell(row=row, column=7, value=dividend.get("net_amount", 0))
            ws.cell(row=row, column=8, value=dividend.get("withholding_tax", 0))

            # Apply number formatting
            for col in [4, 5, 6, 7, 8]:
                ws.cell(row=row, column=col).style = self.styles["number"]

        # Auto-adjust column widths safely
        self._safe_auto_fit_columns(ws)

    def _create_trigger_analysis(self, simulation_result: SimulationResult):
        """Create trigger analysis sheet with all trigger events."""
        ws = self.workbook.create_sheet("Trigger Analysis")

        ws["A1"] = "Trigger Analysis"
        ws["A1"].font = Font(size=16, bold=True, color="2F5597")
        ws.merge_cells("A1:H1")

        ws["A2"] = (
            f"Ticker: {simulation_result.ticker} | Analysis of all trigger events during simulation"
        )
        ws["A2"].font = Font(size=12, italic=True, color="666666")
        ws.merge_cells("A2:H2")

        trigger_analysis = (
            simulation_result.raw_data.get("trigger_analysis", [])
            if simulation_result.raw_data
            else []
        )
        if not trigger_analysis:
            ws["A4"] = "No trigger analysis data available"
            return

        # Headers
        headers = [
            "Timestamp",
            "Price",
            "Anchor Price",
            "Threshold %",
            "Triggered",
            "Side",
            "Executed",
            "Commission",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.style = self.styles["header"]

        # Trigger data
        for row, trigger in enumerate(trigger_analysis, 5):
            ws.cell(row=row, column=1, value=trigger.get("timestamp", ""))
            ws.cell(row=row, column=2, value=trigger.get("price", 0))
            ws.cell(row=row, column=3, value=trigger.get("anchor_price", 0))
            ws.cell(row=row, column=4, value=trigger.get("threshold_pct", 0))
            ws.cell(row=row, column=5, value=trigger.get("triggered", False))
            ws.cell(row=row, column=6, value=trigger.get("side", ""))
            ws.cell(row=row, column=7, value=trigger.get("executed", False))
            ws.cell(row=row, column=8, value=trigger.get("commission", 0))

            # Apply number formatting
            for col in [2, 3, 4, 8]:
                ws.cell(row=row, column=col).style = self.styles["number"]

        # Auto-adjust column widths safely
        self._safe_auto_fit_columns(ws)

    def _create_daily_returns(self, simulation_result: SimulationResult):
        """Create daily returns sheet with portfolio performance."""
        ws = self.workbook.create_sheet("Daily Returns")

        ws["A1"] = "Daily Returns Analysis"
        ws["A1"].font = Font(size=16, bold=True, color="2F5597")
        ws.merge_cells("A1:H1")

        ws["A2"] = (
            f"Ticker: {simulation_result.ticker} | Daily portfolio performance during simulation"
        )
        ws["A2"].font = Font(size=12, italic=True, color="666666")
        ws.merge_cells("A2:H2")

        daily_returns = (
            simulation_result.raw_data.get("daily_returns", [])
            if simulation_result.raw_data
            else []
        )
        if not daily_returns:
            ws["A4"] = "No daily returns data available"
            return

        # Headers
        headers = [
            "Date",
            "Portfolio Value",
            "Cash",
            "Shares",
            "Asset Value",
            "Daily Return %",
            "Cumulative Return %",
            "Price",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.style = self.styles["header"]

        # Daily returns data
        for row, daily in enumerate(daily_returns, 5):
            ws.cell(row=row, column=1, value=daily.get("date", ""))
            ws.cell(row=row, column=2, value=daily.get("portfolio_value", 0))
            ws.cell(row=row, column=3, value=daily.get("cash", 0))
            ws.cell(row=row, column=4, value=daily.get("shares", 0))
            ws.cell(row=row, column=5, value=daily.get("asset_value", 0))
            ws.cell(row=row, column=6, value=daily.get("daily_return_pct", 0))
            ws.cell(row=row, column=7, value=daily.get("cumulative_return_pct", 0))
            ws.cell(row=row, column=8, value=daily.get("price", 0))

            # Apply number formatting
            for col in [2, 3, 4, 5, 8]:
                ws.cell(row=row, column=col).style = self.styles["number"]
            for col in [6, 7]:
                ws.cell(row=row, column=col).style = self.styles["percentage"]

        # Auto-adjust column widths safely
        self._safe_auto_fit_columns(ws)

    def _create_price_data(self, simulation_result: SimulationResult):
        """Create price data sheet with market data."""
        ws = self.workbook.create_sheet("Price Data")

        ws["A1"] = "Market Price Data"
        ws["A1"].font = Font(size=16, bold=True, color="2F5597")
        ws.merge_cells("A1:H1")

        ws["A2"] = f"Ticker: {simulation_result.ticker} | Market data used in simulation"
        ws["A2"].font = Font(size=12, italic=True, color="666666")
        ws.merge_cells("A2:H2")

        price_data = (
            simulation_result.raw_data.get("price_data", []) if simulation_result.raw_data else []
        )
        if not price_data:
            ws["A4"] = "No price data available"
            return

        # Headers
        headers = ["Timestamp", "Price", "Volume", "Market Hours", "Date", "Time"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.style = self.styles["header"]

        # Price data
        for row, price_point in enumerate(price_data, 5):
            timestamp = price_point.get("timestamp", "")
            ws.cell(row=row, column=1, value=timestamp)
            ws.cell(row=row, column=2, value=price_point.get("price", 0))
            ws.cell(row=row, column=3, value=price_point.get("volume", 0))
            ws.cell(row=row, column=4, value=price_point.get("is_market_hours", True))

            # Extract date and time from timestamp
            if timestamp:
                try:
                    from datetime import datetime

                    dt = datetime.fromisoformat(timestamp.replace("Z", "+00:00"))
                    ws.cell(row=row, column=5, value=dt.strftime("%Y-%m-%d"))
                    ws.cell(row=row, column=6, value=dt.strftime("%H:%M:%S"))
                except Exception:
                    ws.cell(row=row, column=5, value="")
                    ws.cell(row=row, column=6, value="")

            # Apply number formatting
            for col in [2, 3]:
                ws.cell(row=row, column=col).style = self.styles["number"]

        # Auto-adjust column widths safely
        self._safe_auto_fit_columns(ws)

    def _create_comprehensive_time_series(self, simulation_result: SimulationResult):
        """Create comprehensive time-series sheet with every data point."""
        ws = self.workbook.create_sheet("Comprehensive Time Series", 1)  # Insert as second sheet

        ws["A1"] = "Comprehensive Time Series Data"
        ws["A1"].font = Font(size=16, bold=True, color="2F5597")
        ws.merge_cells("A1:Z1")

        ws["A2"] = (
            f"Ticker: {simulation_result.ticker} | Every time point with complete portfolio state"
        )
        ws["A2"].font = Font(size=12, italic=True, color="666666")
        ws.merge_cells("A2:Z2")

        time_series_data = (
            simulation_result.raw_data.get("time_series_data", [])
            if simulation_result.raw_data
            else []
        )
        if not time_series_data:
            ws["A4"] = "No time-series data available"
            return

        # Headers for comprehensive time-series data
        headers = [
            "Timestamp",
            "Date",
            "Time",
            "Price",
            "Volume",
            "Market Hours",
            "Anchor Price",
            "Shares",
            "Cash",
            "Asset Value",
            "Total Value",
            "Asset Allocation %",
            "Price Change %",
            "Trigger Threshold %",
            "Triggered",
            "Side",
            "Executed",
            "Commission",
            "Trade Qty",
            "Trade Notional",
            "Dividend Event",
            "Dividend DPS",
            "Dividend Gross",
            "Dividend Net",
            "Dividend Withholding",
            "Execution Error",
            "Reason",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.style = self.styles["header"]

        # Time-series data
        for row, data_point in enumerate(time_series_data, 5):
            ws.cell(row=row, column=1, value=data_point.get("timestamp", ""))
            ws.cell(row=row, column=2, value=data_point.get("date", ""))
            ws.cell(row=row, column=3, value=data_point.get("time", ""))
            ws.cell(row=row, column=4, value=data_point.get("price", 0))
            ws.cell(row=row, column=5, value=data_point.get("volume", 0))
            ws.cell(row=row, column=6, value=data_point.get("is_market_hours", True))
            ws.cell(row=row, column=7, value=data_point.get("anchor_price", 0))
            ws.cell(row=row, column=8, value=data_point.get("shares", 0))
            ws.cell(row=row, column=9, value=data_point.get("cash", 0))
            ws.cell(row=row, column=10, value=data_point.get("asset_value", 0))
            ws.cell(row=row, column=11, value=data_point.get("total_value", 0))
            ws.cell(row=row, column=12, value=data_point.get("asset_allocation_pct", 0))
            ws.cell(row=row, column=13, value=data_point.get("price_change_pct", 0))
            ws.cell(row=row, column=14, value=data_point.get("trigger_threshold_pct", 0))
            ws.cell(row=row, column=15, value=data_point.get("triggered", False))
            ws.cell(row=row, column=16, value=data_point.get("side", ""))
            ws.cell(row=row, column=17, value=data_point.get("executed", False))
            ws.cell(row=row, column=18, value=data_point.get("commission", 0))
            ws.cell(row=row, column=19, value=data_point.get("trade_qty", 0))
            ws.cell(row=row, column=20, value=data_point.get("trade_notional", 0))
            ws.cell(row=row, column=21, value=data_point.get("dividend_event", False))
            ws.cell(row=row, column=22, value=data_point.get("dividend_dps", 0))
            ws.cell(row=row, column=23, value=data_point.get("dividend_gross", 0))
            ws.cell(row=row, column=24, value=data_point.get("dividend_net", 0))
            ws.cell(row=row, column=25, value=data_point.get("dividend_withholding", 0))
            ws.cell(row=row, column=26, value=data_point.get("execution_error", ""))
            ws.cell(row=row, column=27, value=data_point.get("reason", ""))

            # Apply number formatting
            for col in [4, 5, 7, 8, 9, 10, 11, 12, 13, 14, 18, 19, 20, 22, 23, 24, 25]:
                ws.cell(row=row, column=col).style = self.styles["number"]
            for col in [12, 13, 14]:  # Percentage columns
                ws.cell(row=row, column=col).style = self.styles["percentage"]

        # Auto-adjust column widths safely
        self._safe_auto_fit_columns(ws)

    def _create_verbose_timeline(self, simulation_result: SimulationResult, ticker: str):
        """Create verbose timeline sheet with all incremental changes per time step."""
        ws = self.workbook.create_sheet("Verbose Timeline", 1)  # Insert as second sheet

        ws["A1"] = "Verbose Timeline - Incremental Changes Per Time Step"
        ws["A1"].font = Font(size=16, bold=True, color="2F5597")
        ws.merge_cells("A1:AM1")

        ws["A2"] = (
            f"Ticker: {ticker} | Each row represents a single evaluation point with complete state"
        )
        ws["A2"].font = Font(size=12, italic=True, color="666666")
        ws.merge_cells("A2:AM2")

        # Build timeline using VerboseTimelineService
        timeline_service = VerboseTimelineService()

        # Convert simulation result to dict format
        # Handle both domain entity and use case SimulationResult
        if hasattr(simulation_result, "time_series_data"):
            # Use case SimulationResult
            simulation_dict = {
                "time_series_data": getattr(simulation_result, "time_series_data", []) or [],
                "trade_log": getattr(simulation_result, "trade_log", []) or [],
                "dividend_analysis": getattr(simulation_result, "dividend_analysis", {}) or {},
                "trigger_analysis": getattr(simulation_result, "trigger_analysis", []) or [],
            }
        else:
            # Domain entity SimulationResult
            simulation_dict = {
                "time_series_data": (
                    simulation_result.raw_data.get("time_series_data", [])
                    if simulation_result.raw_data
                    else []
                ),
                "trade_log": (
                    simulation_result.raw_data.get("trade_log", [])
                    if simulation_result.raw_data
                    else []
                ),
                "dividend_analysis": (
                    simulation_result.raw_data.get("dividend_analysis", {})
                    if simulation_result.raw_data
                    else {}
                ),
                "trigger_analysis": (
                    simulation_result.raw_data.get("trigger_analysis", [])
                    if simulation_result.raw_data
                    else []
                ),
            }

        timeline_rows = timeline_service.build_timeline_from_simulation(simulation_dict)

        if not timeline_rows:
            ws["A4"] = "No timeline data available"
            return

        # Define all columns in exact order from spec
        headers = [
            "DateTime",
            "Open",
            "High",
            "Low",
            "Close",
            "Volume",
            "AnchorPrice",
            "DividendPercent",
            "DividendValue",
            "PctChangeFromPrev",
            "PctChangeFromAnchor",
            "Quantity",
            "PositionValue",
            "Cash",
            "TotalPortfolioValue",
            "DeltaTotalValue",
            "PctDeltaTotalValue",
            "PctStockChangeFromBaseline",
            "PctPortfolioChangeFromBaseline",
            "TriggerThresholdUp",
            "TriggerThresholdDown",
            "TriggerSignal",
            "TriggerReason",
            "GuardrailMinStockPct",
            "GuardrailMaxStockPct",
            "GuardrailMaxTradePctOfPosition",
            "GuardrailAllowed",
            "GuardrailReason",
            "Action",
            "ActionReason",
            "TradeSide",
            "TradeQuantity",
            "TradePrice",
            "TradeNotional",
            "TradeCommission",
            "TradeCommissionRateEffective",
            "DividendCashThisLine",
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.style = self.styles["header"]

        # Write data rows
        for row_idx, row_data in enumerate(timeline_rows, 5):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, None)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Apply appropriate formatting
                if header in ["Open", "High", "Low", "Close", "AnchorPrice", "TradePrice"]:
                    cell.style = self.styles["number"]
                    cell.number_format = "#,##0.00"
                elif header in [
                    "Volume",
                    "Quantity",
                    "TradeQuantity",
                    "PositionValue",
                    "Cash",
                    "TotalPortfolioValue",
                    "DeltaTotalValue",
                    "TradeNotional",
                    "TradeCommission",
                    "DividendValue",
                    "DividendCashThisLine",
                ]:
                    cell.style = self.styles["number"]
                    cell.number_format = "#,##0.00"
                elif header in [
                    "DividendPercent",
                    "PctChangeFromPrev",
                    "PctChangeFromAnchor",
                    "PctDeltaTotalValue",
                    "PctStockChangeFromBaseline",
                    "PctPortfolioChangeFromBaseline",
                    "TriggerThresholdUp",
                    "TriggerThresholdDown",
                    "GuardrailMinStockPct",
                    "GuardrailMaxStockPct",
                    "GuardrailMaxTradePctOfPosition",
                    "TradeCommissionRateEffective",
                ]:
                    cell.style = self.styles["percentage"]
                    cell.number_format = "0.00"
                elif header == "DateTime":
                    # Keep as string for ISO format
                    cell.style = self.styles["data"]
                else:
                    cell.style = self.styles["data"]

        # Auto-adjust column widths safely
        self._safe_auto_fit_columns(ws)

        # Freeze header row
        ws.freeze_panes = "A5"

    def create_activity_log_report(
        self,
        events: List[Dict[str, Any]],
        position_id: str,
        ticker: str = "UNKNOWN",
    ) -> bytes:
        """Create an activity log report from events."""
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)

        # Define styles
        self._define_styles()

        # Create activity log sheet
        ws = self.workbook.create_sheet("Activity Log", 0)

        # Title
        ws["A1"] = f"Activity Log - {ticker} (Position: {position_id})"
        ws["A1"].font = Font(size=18, bold=True, color="2F5597")
        ws.merge_cells("A1:H1")

        # Generation info
        ws["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}"
        ws["A2"].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells("A2:H2")

        if not events:
            ws["A4"] = "No events found for this position"
            excel_buffer = io.BytesIO()
            self.workbook.save(excel_buffer)
            excel_buffer.seek(0)
            return excel_buffer.getvalue()

        # Headers
        headers = ["Timestamp", "Type", "Message", "Inputs", "Outputs"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.style = self.styles["header"]

        # Event data
        for row, event in enumerate(events, 5):
            ws.cell(row=row, column=1, value=event.get("ts", ""))
            ws.cell(row=row, column=2, value=event.get("type", ""))
            ws.cell(row=row, column=3, value=event.get("message", ""))
            ws.cell(
                row=row,
                column=4,
                value=json.dumps(event.get("inputs", {}), indent=2) if event.get("inputs") else "",
            )
            ws.cell(
                row=row,
                column=5,
                value=(
                    json.dumps(event.get("outputs", {}), indent=2) if event.get("outputs") else ""
                ),
            )

        # Auto-adjust column widths safely
        self._safe_auto_fit_columns(ws)

        # Save to bytes
        excel_buffer = io.BytesIO()
        self.workbook.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()

    def create_dividend_report(
        self,
        receivables: List[Dict[str, Any]],
        upcoming_dividends: List[Dict[str, Any]],
        summary: Dict[str, Any],
    ) -> bytes:
        """Create a dividend report with Receivables and Upcoming sheets."""
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)
        self._define_styles()

        ticker = summary.get("ticker", "UNKNOWN")
        position_id = summary.get("position_id", "")

        # --- Sheet 1: Receivables ---
        ws = self.workbook.create_sheet("Receivables", 0)
        ws["A1"] = f"Dividend Receivables - {ticker}"
        ws["A1"].font = Font(size=18, bold=True, color="2F5597")
        ws.merge_cells("A1:H1")
        ws["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}"
        ws["A2"].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells("A2:H2")

        # Summary row
        ws["A3"] = f"Total Pending: {summary.get('total_pending', 0)} | Total Paid: {summary.get('total_paid', 0)}"
        ws["A3"].font = Font(size=10, color="333333")
        ws.merge_cells("A3:H3")

        headers = [
            "Status", "Ex-Date", "Pay-Date", "Shares at Record",
            "Gross Amount", "Withholding Tax", "Net Amount", "Paid At",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.style = self.styles["header"]

        for row_idx, rec in enumerate(receivables, 6):
            ws.cell(row=row_idx, column=1, value=rec.get("status", ""))
            ws.cell(row=row_idx, column=2, value=rec.get("ex_date", ""))
            ws.cell(row=row_idx, column=3, value=rec.get("pay_date", ""))
            ws.cell(row=row_idx, column=4, value=rec.get("shares_at_record", ""))
            ws.cell(row=row_idx, column=5, value=rec.get("gross_amount", ""))
            ws.cell(row=row_idx, column=6, value=rec.get("withholding_tax", ""))
            ws.cell(row=row_idx, column=7, value=rec.get("net_amount", ""))
            ws.cell(row=row_idx, column=8, value=rec.get("paid_at", ""))

        self._safe_auto_fit_columns(ws)

        # --- Sheet 2: Upcoming Dividends ---
        ws2 = self.workbook.create_sheet("Upcoming Dividends")
        ws2["A1"] = f"Upcoming Dividends - {ticker}"
        ws2["A1"].font = Font(size=18, bold=True, color="2F5597")
        ws2.merge_cells("A1:D1")
        ws2["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}"
        ws2["A2"].font = Font(size=10, italic=True, color="666666")
        ws2.merge_cells("A2:D2")

        upcoming_headers = ["Ex-Date", "Pay-Date", "DPS", "Currency"]
        for col, header in enumerate(upcoming_headers, 1):
            cell = ws2.cell(row=4, column=col, value=header)
            cell.style = self.styles["header"]

        for row_idx, div in enumerate(upcoming_dividends, 5):
            ws2.cell(row=row_idx, column=1, value=div.get("ex_date", ""))
            ws2.cell(row=row_idx, column=2, value=div.get("pay_date", ""))
            ws2.cell(row=row_idx, column=3, value=div.get("dps", ""))
            ws2.cell(row=row_idx, column=4, value=div.get("currency", "USD"))

        self._safe_auto_fit_columns(ws2)

        # Save to bytes
        excel_buffer = io.BytesIO()
        self.workbook.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()
