# backend/app/routes/explainability.py
"""
Explainability API endpoints.

Provides endpoints for the Explainability Table feature, supporting both
live trading and simulation data sources. Includes unified trade tracking
with order/trade enrichment, pagination, and filtering.
"""

from typing import Optional, List
from datetime import datetime
from uuid import UUID
from io import BytesIO

from fastapi import APIRouter, HTTPException, Query, Depends
from fastapi.responses import StreamingResponse

from app.di import container
from application.services.explainability_timeline_service import ExplainabilityTimelineService


# Create router for live position explainability
live_router = APIRouter(
    prefix="/v1/tenants/{tenant_id}/portfolios/{portfolio_id}/positions/{position_id}",
    tags=["explainability"],
)

# Create router for simulation explainability
simulation_router = APIRouter(
    prefix="/v1/simulations",
    tags=["explainability"],
)


def get_explainability_service() -> ExplainabilityTimelineService:
    """Get explainability timeline service instance with order/trade repos."""
    return ExplainabilityTimelineService(
        orders_repo=container.orders,
        trades_repo=container.trades,
    )


def _parse_date(date_str: Optional[str], param_name: str) -> Optional[datetime]:
    """Parse a date string parameter, raising HTTPException on invalid format."""
    if not date_str:
        return None
    try:
        return datetime.fromisoformat(date_str)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid {param_name} format: {date_str}")


def _parse_csv_filter(value: Optional[str]) -> Optional[List[str]]:
    """Parse a comma-separated filter string into a list."""
    if not value:
        return None
    return [item.strip() for item in value.split(",") if item.strip()]


@live_router.get("/explainability")
def get_live_explainability(
    tenant_id: str,
    portfolio_id: str,
    position_id: str,
    start_date: Optional[str] = Query(None, description="Start date filter (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date filter (YYYY-MM-DD)"),
    action: Optional[str] = Query(None, description="Action filter (comma-separated: BUY,SELL,HOLD,SKIP)"),
    order_status: Optional[str] = Query(None, description="Order status filter (comma-separated: filled,rejected,cancelled,pending,working)"),
    aggregation: str = Query("daily", description="Aggregation mode: 'daily' or 'all'"),
    offset: int = Query(0, ge=0, description="Number of rows to skip for pagination"),
    limit: int = Query(500, ge=1, le=2000, description="Maximum number of rows to return"),
    service: ExplainabilityTimelineService = Depends(get_explainability_service),
) -> dict:
    """
    Get explainability timeline for a live position.

    Returns a unified view of how the trading algorithm behaved at each timestamp,
    with all the data needed to understand why decisions were made. Includes
    order lifecycle and trade execution details.

    Supports pagination via offset/limit and filtering by action type, order status,
    and date range.
    """
    try:
        # Validate timeline repository is available
        if not hasattr(container, "evaluation_timeline"):
            raise HTTPException(status_code=501, detail="Timeline repository not available")

        # Parse filters
        start_dt = _parse_date(start_date, "start_date")
        end_dt = _parse_date(end_date, "end_date")
        actions_list = _parse_csv_filter(action)
        if actions_list:
            actions_list = [a.upper() for a in actions_list]
        order_statuses = _parse_csv_filter(order_status)

        # Get position info for metadata
        position = container.positions.get(
            tenant_id=tenant_id, portfolio_id=portfolio_id, position_id=position_id
        )
        symbol = position.asset_symbol if position else None

        # Fetch timeline data from repository
        timeline_records = container.evaluation_timeline.list_by_position(
            tenant_id=tenant_id,
            portfolio_id=portfolio_id,
            position_id=position_id,
            mode="LIVE",
            limit=limit * 10,  # Fetch more to allow for filtering
        )

        # Convert to explainability rows
        rows = service.build_from_live_timeline(
            evaluation_records=timeline_records,
            position_id=position_id,
            portfolio_id=portfolio_id,
            symbol=symbol,
        )

        # Enrich with order/trade data (populates Groups 6-7)
        rows = service.enrich_with_orders(rows)

        # Build timeline with filtering, aggregation, and pagination
        timeline = service.build_timeline(
            rows=rows,
            start_date=start_dt,
            end_date=end_dt,
            actions=actions_list,
            order_statuses=order_statuses,
            aggregation=aggregation,
            offset=offset,
            limit=limit,
            position_id=position_id,
            portfolio_id=portfolio_id,
            symbol=symbol,
            mode="LIVE",
        )

        return timeline.to_dict()

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error getting explainability data: {str(e)}")


@live_router.get("/explainability/export")
def export_live_explainability(
    tenant_id: str,
    portfolio_id: str,
    position_id: str,
    start_date: Optional[str] = Query(None, description="Start date filter (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date filter (YYYY-MM-DD)"),
    action: Optional[str] = Query(None, description="Action filter (comma-separated: BUY,SELL,HOLD,SKIP)"),
    order_status: Optional[str] = Query(None, description="Order status filter (comma-separated)"),
    aggregation: str = Query("daily", description="Aggregation mode: 'daily' or 'all'"),
    service: ExplainabilityTimelineService = Depends(get_explainability_service),
):
    """
    Export explainability timeline to Excel.

    Returns an Excel file with the explainability timeline data including
    order and trade details.
    """
    try:
        # Validate timeline repository is available
        if not hasattr(container, "evaluation_timeline"):
            raise HTTPException(status_code=501, detail="Timeline repository not available")

        # Parse filters
        start_dt = _parse_date(start_date, "start_date")
        end_dt = _parse_date(end_date, "end_date")
        actions_list = _parse_csv_filter(action)
        if actions_list:
            actions_list = [a.upper() for a in actions_list]
        order_statuses = _parse_csv_filter(order_status)

        # Get position info
        position = container.positions.get(
            tenant_id=tenant_id, portfolio_id=portfolio_id, position_id=position_id
        )
        symbol = position.asset_symbol if position else "unknown"

        # Fetch all timeline data
        timeline_records = container.evaluation_timeline.list_by_position(
            tenant_id=tenant_id,
            portfolio_id=portfolio_id,
            position_id=position_id,
            mode="LIVE",
            limit=10000,
        )

        # Convert to explainability rows
        rows = service.build_from_live_timeline(
            evaluation_records=timeline_records,
            position_id=position_id,
            portfolio_id=portfolio_id,
            symbol=symbol,
        )

        # Enrich with order/trade data
        rows = service.enrich_with_orders(rows)

        # Build timeline with filtering and aggregation
        timeline = service.build_timeline(
            rows=rows,
            start_date=start_dt,
            end_date=end_dt,
            actions=actions_list,
            order_statuses=order_statuses,
            aggregation=aggregation,
            limit=10000,
            position_id=position_id,
            portfolio_id=portfolio_id,
            symbol=symbol,
            mode="LIVE",
        )

        # Generate Excel
        excel_data = _generate_excel(timeline.rows, symbol, "LIVE")

        filename = f"explainability_{symbol}_{position_id[:8]}_{datetime.now().strftime('%Y%m%d')}.xlsx"

        return StreamingResponse(
            iter([excel_data]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error exporting explainability data: {str(e)}")


@simulation_router.get("/{simulation_id}/explainability")
def get_simulation_explainability(
    simulation_id: str,
    start_date: Optional[str] = Query(None, description="Start date filter (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date filter (YYYY-MM-DD)"),
    action: Optional[str] = Query(None, description="Action filter (comma-separated: BUY,SELL,HOLD,SKIP)"),
    aggregation: str = Query("daily", description="Aggregation mode: 'daily' or 'all'"),
    offset: int = Query(0, ge=0, description="Number of rows to skip for pagination"),
    limit: int = Query(500, ge=1, le=2000, description="Maximum number of rows to return"),
    service: ExplainabilityTimelineService = Depends(get_explainability_service),
) -> dict:
    """
    Get explainability timeline for a simulation.

    Returns a unified view of how the trading algorithm behaved at each timestamp
    during the simulation. Supports pagination via offset/limit.
    """
    try:
        # Get simulation from repository
        simulation = container.simulation.get_simulation_result(UUID(simulation_id))
        if not simulation:
            raise HTTPException(status_code=404, detail="Simulation not found")

        # Parse filters
        start_dt = _parse_date(start_date, "start_date")
        end_dt = _parse_date(end_date, "end_date")
        actions_list = _parse_csv_filter(action)
        if actions_list:
            actions_list = [a.upper() for a in actions_list]

        # Build simulation result dict
        raw_data = simulation.raw_data or {}
        simulation_result = {
            "id": str(simulation.id),
            "ticker": simulation.ticker,
            "time_series_data": raw_data.get("time_series_data", []),
            "trade_log": simulation.trade_log or [],
        }

        # Convert to explainability rows
        rows = service.build_from_simulation(simulation_result)

        # Build timeline with filtering, aggregation, and pagination
        timeline = service.build_timeline(
            rows=rows,
            start_date=start_dt,
            end_date=end_dt,
            actions=actions_list,
            aggregation=aggregation,
            offset=offset,
            limit=limit,
            simulation_run_id=simulation_id,
            symbol=simulation.ticker,
            mode="SIMULATION",
        )

        return timeline.to_dict()

    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid simulation ID format")
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error getting explainability data: {str(e)}")


@simulation_router.get("/{simulation_id}/explainability/export")
def export_simulation_explainability(
    simulation_id: str,
    start_date: Optional[str] = Query(None, description="Start date filter (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date filter (YYYY-MM-DD)"),
    action: Optional[str] = Query(None, description="Action filter (comma-separated: BUY,SELL,HOLD,SKIP)"),
    aggregation: str = Query("daily", description="Aggregation mode: 'daily' or 'all'"),
    service: ExplainabilityTimelineService = Depends(get_explainability_service),
):
    """
    Export simulation explainability timeline to Excel.
    """
    try:
        # Get simulation from repository
        simulation = container.simulation.get_simulation_result(UUID(simulation_id))
        if not simulation:
            raise HTTPException(status_code=404, detail="Simulation not found")

        # Parse filters
        start_dt = _parse_date(start_date, "start_date")
        end_dt = _parse_date(end_date, "end_date")
        actions_list = _parse_csv_filter(action)
        if actions_list:
            actions_list = [a.upper() for a in actions_list]

        # Build simulation result dict
        raw_data = simulation.raw_data or {}
        simulation_result = {
            "id": str(simulation.id),
            "ticker": simulation.ticker,
            "time_series_data": raw_data.get("time_series_data", []),
            "trade_log": simulation.trade_log or [],
        }

        # Convert to explainability rows
        rows = service.build_from_simulation(simulation_result)

        # Build timeline
        timeline = service.build_timeline(
            rows=rows,
            start_date=start_dt,
            end_date=end_dt,
            actions=actions_list,
            aggregation=aggregation,
            limit=10000,
            simulation_run_id=simulation_id,
            symbol=simulation.ticker,
            mode="SIMULATION",
        )

        # Generate Excel
        excel_data = _generate_excel(timeline.rows, simulation.ticker, "SIMULATION")

        filename = f"explainability_{simulation.ticker}_{simulation_id[:8]}_{datetime.now().strftime('%Y%m%d')}.xlsx"

        return StreamingResponse(
            iter([excel_data]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid simulation ID format")
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error exporting explainability data: {str(e)}")


def _generate_excel(rows: List, symbol: str, mode: str) -> bytes:
    """Generate Excel file from explainability rows with all column groups."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed. Run: pip install openpyxl"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Explainability"

    # Define columns grouped by category
    columns = [
        # Group 1: Time & Identity
        ("Timestamp", "timestamp"),
        ("Date", "date"),
        ("Trace ID", "trace_id"),
        # Group 2: Market Data
        ("Price", "price"),
        ("Open", "open"),
        ("High", "high"),
        ("Low", "low"),
        ("Close", "close"),
        ("Volume", "volume"),
        # Group 3: Trigger Evaluation
        ("Anchor", "anchor_price"),
        ("Delta %", "delta_pct"),
        ("Trigger Up %", "trigger_up_threshold"),
        ("Trigger Down %", "trigger_down_threshold"),
        ("Trigger Fired", "trigger_fired"),
        ("Trigger Direction", "trigger_direction"),
        # Group 4: Guardrails
        ("Stock % Before", "current_stock_pct"),
        ("Min %", "min_stock_pct"),
        ("Max %", "max_stock_pct"),
        ("Guardrail OK", "guardrail_allowed"),
        ("Block Reason", "guardrail_block_reason"),
        # Group 5: Action Decision
        ("Action", "action"),
        ("Reason", "action_reason"),
        ("Intended Qty", "intended_qty"),
        ("Intended Value", "intended_value"),
        # Group 6: Order Status
        ("Order ID", "order_id"),
        ("Order Status", "order_status"),
        ("Broker Order ID", "broker_order_id"),
        ("Broker Status", "broker_status"),
        # Group 7: Execution Details
        ("Exec Price", "execution_price"),
        ("Exec Qty", "execution_qty"),
        ("Exec Value", "execution_value"),
        ("Commission", "commission"),
        ("Exec Status", "execution_status"),
        # Group 8: Position Impact
        ("Qty Before", "qty_before"),
        ("Qty After", "qty_after"),
        ("Cash Before", "cash_before"),
        ("Cash After", "cash_after"),
        ("Stock Val Before", "stock_value_before"),
        ("Stock Val After", "stock_value_after"),
        ("Total Val Before", "total_value_before"),
        ("Total Val After", "total_value_after"),
        ("Stock % Before", "stock_pct_before"),
        ("Stock % After", "stock_pct_after"),
        # Group 9: Dividends
        ("Dividend", "dividend_declared"),
        # Group 10: Anchor Tracking
        ("Anchor Reset", "anchor_reset"),
        ("Old Anchor", "anchor_old_value"),
        ("Reset Reason", "anchor_reset_reason"),
    ]

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Write headers
    for col_idx, (label, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Row fill styles
    buy_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    sell_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    # Numeric columns for rounding
    money_columns = {
        "price", "open", "high", "low", "close", "anchor_price",
        "stock_value", "cash", "total_value", "execution_price", "commission",
        "intended_value", "execution_value",
        "cash_before", "cash_after", "stock_value_before", "stock_value_after",
        "total_value_before", "total_value_after",
    }
    pct_columns = {
        "delta_pct", "stock_pct", "min_stock_pct", "max_stock_pct",
        "current_stock_pct", "stock_pct_before", "stock_pct_after",
        "trigger_up_threshold", "trigger_down_threshold",
    }
    bool_columns = {"guardrail_allowed", "trigger_fired", "dividend_declared", "anchor_reset"}

    # Write data
    for row_idx, row in enumerate(rows, 2):
        row_dict = row.to_dict()
        action = row_dict.get("action", "")

        for col_idx, (_, key) in enumerate(columns, 1):
            value = row_dict.get(key)

            # Format values
            if value is None:
                value = ""
            elif key in money_columns:
                if value:
                    value = round(float(value), 2)
            elif key in pct_columns:
                if value:
                    value = round(float(value), 2)
            elif key in bool_columns:
                value = "Yes" if value else "No"

            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Apply row highlighting
            if action == "BUY":
                cell.fill = buy_fill
            elif action == "SELL":
                cell.fill = sell_fill

    # Auto-size columns
    for col_idx, (label, _) in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(label)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
